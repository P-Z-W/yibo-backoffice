"""Customer and operating-record module integration tests."""

from datetime import date, datetime
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, select

from app.api.v1 import operation_records as operation_records_api
from app.core.config import settings
from app.db.session import SessionLocal
from app.main import app
from app.models.analytics import MetricDefinition, MonthlyMetric
from app.models.operation_record import (
    CustomerChangeRecord,
    CustomerServiceRecord,
    ShortVideoRecord,
    ValueAddedRecord,
)
from app.models.user import AuditLog

TEST_MONTH = "2033-04"
TEST_DATE = date(2033, 4, 1)
TEST_SOURCE_TEAM_ID = 90909091


def login(client: TestClient) -> None:
    response = client.post(
        f"{settings.api_prefix}/auth/login",
        json={
            "username": settings.initial_admin_username,
            "password": settings.initial_admin_password,
        },
    )
    assert response.status_code == 200


def cleanup() -> None:
    with SessionLocal() as db:
        db.execute(
            delete(CustomerChangeRecord).where(
                CustomerChangeRecord.source_team_id == TEST_SOURCE_TEAM_ID
            )
        )
        resource_ids: list[str] = []
        for dataset, model in (
            ("customer_changes", CustomerChangeRecord),
            ("value_added", ValueAddedRecord),
            ("service_issues", CustomerServiceRecord),
            ("short_video", ShortVideoRecord),
        ):
            ids = list(db.scalars(select(model.id).where(model.month == TEST_DATE)))
            resource_ids.extend(
                f"operation_record:{dataset}:{record_id}" for record_id in ids
            )
            db.execute(delete(model).where(model.month == TEST_DATE))
        if resource_ids:
            db.execute(delete(AuditLog).where(AuditLog.resource.in_(resource_ids)))
        db.execute(
            delete(AuditLog).where(
                AuditLog.resource.in_(
                    [
                        f"operation_records:{dataset}:{TEST_MONTH}"
                        for dataset in (
                            "customer_changes",
                            "value_added",
                            "service_issues",
                            "short_video",
                        )
                    ]
                )
            )
        )
        db.execute(
            delete(AuditLog).where(AuditLog.resource == "customer_source:team_source")
        )
        metric_ids = list(
            db.scalars(
                select(MetricDefinition.id).where(
                    MetricDefinition.code.in_(
                        ("new_customers", "lost_customers", "prospective_customers")
                    )
                )
            )
        )
        db.execute(
            delete(MonthlyMetric).where(
                MonthlyMetric.month == TEST_DATE,
                MonthlyMetric.metric_id.in_(metric_ids),
            )
        )
        db.commit()


def workbook_bytes(headers: list[object], values: list[object]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    sheet.append(values)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_operation_modules_excel_and_analytics_integration() -> None:
    cleanup()
    try:
        with TestClient(app) as client:
            login(client)
            samples = {
                "customer_changes": [
                    None,
                    "新进",
                    datetime(2033, 4, 15, 10, 30, 0),
                    "测试客户A",
                    "短视频",
                    2,
                    "测试导入",
                ],
                "value_added": [
                    None,
                    "T001",
                    "测试团队",
                    "VAS001",
                    "商品贴标",
                    "仓内加工",
                    12,
                ],
                "service_issues": [
                    None,
                    "测试团队",
                    "物流问题",
                    "客户反馈物流时效异常",
                    "中转延误",
                    "物流商",
                    "调整承运渠道",
                    "整改中",
                ],
                "short_video": [None, 5, "产品口播", "测试运营", "本月计划"],
            }
            for dataset, values in samples.items():
                template = client.get(
                    f"{settings.api_prefix}/operation-records/{dataset}/template"
                )
                assert template.status_code == 200
                template_book = load_workbook(BytesIO(template.content))
                headers = [cell.value for cell in template_book.active[1]]
                assert headers[0] == "记录ID"
                imported = client.post(
                    f"{settings.api_prefix}/operation-records/{dataset}/import",
                    params={"month": TEST_MONTH},
                    files={
                        "file": (
                            f"{dataset}.xlsx",
                            workbook_bytes(headers, values),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    },
                )
                assert imported.status_code == 200, imported.text
                assert imported.json()["created_count"] == 1

                listed = client.get(
                    f"{settings.api_prefix}/operation-records/{dataset}",
                    params={"month": TEST_MONTH, "page": 1, "size": 50},
                )
                assert listed.status_code == 200
                assert listed.json()["total"] == 1
                if dataset == "value_added":
                    assert listed.json()["summary"] == {
                        "total": 1,
                        "quantity": 12,
                        "teams": 1,
                        "services": 1,
                    }

                exported = client.get(
                    f"{settings.api_prefix}/operation-records/{dataset}/export",
                    params={"month": TEST_MONTH},
                )
                assert exported.status_code == 200
                export_book = load_workbook(BytesIO(exported.content), data_only=True)
                assert export_book.active["A2"].value == listed.json()["records"][0]["id"]

                details = client.get(
                    f"{settings.api_prefix}/analytics/details/{dataset}",
                    params={"month": TEST_MONTH},
                )
                assert details.status_code == 200
                assert details.json()["total"] == 1
                assert details.json()["batches"][0]["mode"] == "system"

            analytics = client.get(
                f"{settings.api_prefix}/analytics", params={"month": TEST_MONTH}
            )
            assert analytics.status_code == 200
            metric = next(
                item
                for item in analytics.json()["metrics"]
                if item["code"] == "new_customers"
            )
            assert metric["value"] == 2
            assert metric["source_type"] == "operation_module"
            completion = {
                item["code"]: item for item in analytics.json()["completion"]["items"]
            }
            for dataset in samples:
                assert completion[dataset]["state"] == "system"
                assert completion[dataset]["label"] == "模块取数"

            short_rows = client.get(
                f"{settings.api_prefix}/operation-records/short_video",
                params={"month": TEST_MONTH, "page": 1, "size": 50},
            ).json()["records"]
            short_id = short_rows[0]["id"]
            updated = client.put(
                f"{settings.api_prefix}/operation-records/short_video/{short_id}",
                json={
                    "month": TEST_MONTH,
                    "video_count": 6,
                    "video_type": "产品口播",
                    "owner": "测试运营",
                    "note": "更新后的计划",
                },
            )
            assert updated.status_code == 200
            assert updated.json()["video_count"] == 6
            deleted = client.delete(
                f"{settings.api_prefix}/operation-records/short_video/{short_id}"
            )
            assert deleted.status_code == 200
    finally:
        cleanup()


def test_customer_source_preview(monkeypatch) -> None:
    cleanup()
    def fake_preview() -> dict[str, object]:
        return {
            "source": "yibo.team_source",
            "filters": {
                "cooperation_type": 20,
                "stock_send_price_lt": 99900,
                "stock_send_price_null_included": True,
            },
            "total": 1,
            "rows": [
                {
                    "team_id": TEST_SOURCE_TEAM_ID,
                    "team_name": "测试团队",
                    "created_time": "2026-08-29 14:01:49",
                    "cooperation_type": 20,
                    "stock_send_price": 320,
                    "viewable": True,
                }
            ],
        }

    monkeypatch.setattr(operation_records_api, "fetch_customer_source_preview", fake_preview)
    with TestClient(app) as client:
        login(client)
        response = client.get(
            f"{settings.api_prefix}/operation-records/customer_changes/source-preview"
        )
        assert response.status_code == 200
        payload = response.json()
        assert payload["filters"] == {
            "cooperation_type": 20,
            "stock_send_price_lt": 99900,
            "stock_send_price_null_included": True,
        }
        assert payload["total"] == 1
        assert payload["registered_total"] == 0
        assert payload["archived_total"] == 0
        assert payload["pending_total"] == 1
        assert payload["rows"][0] == {
        "team_id": TEST_SOURCE_TEAM_ID,
            "team_name": "测试团队",
            "created_time": "2026-08-29 14:01:49",
            "cooperation_type": 20,
            "stock_send_price": 320,
            "viewable": True,
            "registered": False,
            "archived": False,
        }


def test_customer_source_sync_uses_full_timestamp(monkeypatch) -> None:
    cleanup()

    def fake_preview() -> dict[str, object]:
        return {
            "source": "yibo.team_source",
            "filters": {
                "cooperation_type": 20,
                "stock_send_price_lt": 99900,
                "stock_send_price_null_included": True,
            },
            "total": 1,
            "rows": [
                {
                    "team_id": TEST_SOURCE_TEAM_ID,
                    "team_name": "完整时间测试客户",
                    "created_time": "2033-04-30 23:59:59",
                    "cooperation_type": 20,
                    "stock_send_price": 320,
                    "viewable": True,
                }
            ],
        }

    monkeypatch.setattr(operation_records_api, "fetch_customer_source_preview", fake_preview)
    try:
        with TestClient(app) as client:
            login(client)
            first = client.post(
                f"{settings.api_prefix}/operation-records/customer_changes/source-sync"
            )
            assert first.status_code == 200, first.text
            assert first.json()["created_count"] == 1
            assert first.json()["affected_months"] == [TEST_MONTH]

            second = client.post(
                f"{settings.api_prefix}/operation-records/customer_changes/source-sync"
            )
            assert second.status_code == 200, second.text
            assert second.json()["created_count"] == 0
            assert second.json()["skipped_existing"] == 1

            april = client.get(
                f"{settings.api_prefix}/operation-records/customer_changes",
                params={"month": TEST_MONTH, "page": 1, "size": 50},
            )
            assert april.status_code == 200
            assert april.json()["total"] == 1
            record = april.json()["records"][0]
            assert record["source_team_id"] == TEST_SOURCE_TEAM_ID
            assert record["occurred_at"] == "2033-04-30T23:59:59"
            assert record["month"] == ""
            assert april.json()["summary"]["new"] == 1

            details = client.get(
                f"{settings.api_prefix}/analytics/details/customer_changes",
                params={"month": TEST_MONTH, "page": 1, "size": 50},
            )
            assert details.status_code == 200
            assert details.json()["columns"][1] == "发生时间"
            assert details.json()["rows"][0]["values"]["发生时间"] == "2033-04-30T23:59:59"

            analytics = client.get(
                f"{settings.api_prefix}/analytics", params={"month": TEST_MONTH}
            )
            assert analytics.status_code == 200
            new_customers = next(
                item for item in analytics.json()["metrics"] if item["code"] == "new_customers"
            )
            assert new_customers["value"] == 1

            may = client.get(
                f"{settings.api_prefix}/operation-records/customer_changes",
                params={"month": "2033-05", "page": 1, "size": 50},
            )
            assert may.status_code == 200
            assert may.json()["total"] == 0
    finally:
        cleanup()
