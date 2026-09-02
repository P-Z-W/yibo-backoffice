"""Supplier management and operating-analysis integration tests."""

from datetime import date
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.core.config import settings
from app.db.session import SessionLocal
from app.main import app
from app.models.analytics import MetricDefinition, MonthlyMetric
from app.models.supplier import Supplier, SupplierChange
from app.models.user import AuditLog

TEST_NAME = "供应商模块自动测试"
TEST_NORMALIZED_NAME = TEST_NAME.casefold()
TEST_EXCEL_NAME = "供应商Excel自动测试"
TEST_BATCH_EXCEL_NAME = "供应商批量上传自动测试"
TEST_NORMALIZED_NAMES = {
    TEST_NORMALIZED_NAME,
    TEST_EXCEL_NAME.casefold(),
    TEST_BATCH_EXCEL_NAME.casefold(),
}
TEST_MONTH = "2032-03"


def login(client: TestClient) -> None:
    response = client.post(
        f"{settings.api_prefix}/auth/login",
        json={
            "username": settings.initial_admin_username,
            "password": settings.initial_admin_password,
        },
    )
    assert response.status_code == 200


def cleanup() -> None:
    selected = date(2032, 3, 1)
    with SessionLocal() as db:
        supplier_ids = list(
            db.scalars(
                select(Supplier.id).where(
                    Supplier.normalized_name.in_(TEST_NORMALIZED_NAMES)
                )
            )
        )
        db.execute(
            delete(AuditLog).where(AuditLog.resource == f"suppliers:{TEST_MONTH}")
        )
        if supplier_ids:
            db.execute(
                delete(AuditLog).where(
                    AuditLog.resource.in_([f"supplier:{item}" for item in supplier_ids])
                )
            )
            db.execute(
                delete(SupplierChange).where(
                    SupplierChange.supplier_id.in_(supplier_ids)
                )
            )
            db.execute(delete(Supplier).where(Supplier.id.in_(supplier_ids)))
        metric_id = db.scalar(
            select(MetricDefinition.id).where(
                MetricDefinition.code == "supplier_change"
            )
        )
        if metric_id is not None:
            db.execute(
                delete(MonthlyMetric).where(
                    MonthlyMetric.metric_id == metric_id,
                    MonthlyMetric.month == selected,
                )
            )
        db.commit()


def test_supplier_lifecycle_feeds_monthly_analytics() -> None:
    cleanup()
    try:
        with TestClient(app) as client:
            login(client)
            created = client.post(
                f"{settings.api_prefix}/suppliers",
                json={
                    "name": TEST_NAME,
                    "contact_name": "张供应",
                    "contact_phone": "13800000000",
                    "address": "测试地址",
                    "cooperation_start_date": "2032-03-02",
                    "product_types": "纸箱,胶带",
                    "note": "初始合作",
                    "change_month": TEST_MONTH,
                    "change_note": "首次建档",
                },
            )
            assert created.status_code == 201
            supplier_id = created.json()["id"]

            listed = client.get(
                f"{settings.api_prefix}/suppliers",
                params={"month": TEST_MONTH, "keyword": TEST_NAME},
            )
            assert listed.status_code == 200
            assert listed.json()["summary"]["month_added"] == 1
            assert listed.json()["summary"]["month_changed"] == 1

            updated = client.put(
                f"{settings.api_prefix}/suppliers/{supplier_id}",
                json={
                    "name": TEST_NAME,
                    "contact_name": "李供应",
                    "contact_phone": "13900000000",
                    "address": "测试地址二期",
                    "cooperation_start_date": "2032-03-02",
                    "product_types": "纸箱,胶带,气泡膜",
                    "note": "扩充品类",
                    "change_month": TEST_MONTH,
                    "change_note": "联系人和品类调整",
                },
            )
            assert updated.status_code == 200
            assert updated.json()["contact_name"] == "李供应"

            changes = client.get(
                f"{settings.api_prefix}/suppliers/changes",
                params={"month": TEST_MONTH},
            )
            assert changes.status_code == 200
            assert [item["change_type"] for item in changes.json()[:2]] == [
                "updated",
                "created",
            ]

            details = client.get(
                f"{settings.api_prefix}/analytics/details/supplier_changes",
                params={"month": TEST_MONTH},
            )
            assert details.status_code == 200
            assert details.json()["total"] == 1
            assert details.json()["summary"]["supplier_change"] == 1
            assert details.json()["rows"][0]["values"]["供应商名称"] == TEST_NAME
            assert details.json()["rows"][0]["values"]["供应商联系人"] == "李供应"

            analytics = client.get(
                f"{settings.api_prefix}/analytics", params={"month": TEST_MONTH}
            )
            assert analytics.status_code == 200
            supplier_completion = next(
                item
                for item in analytics.json()["completion"]["items"]
                if item["code"] == "supplier_changes"
            )
            assert supplier_completion["state"] == "system"
            assert supplier_completion["label"] == "模块取数"
            assert supplier_completion["source_name"] == "供应商管理模块"
            assert supplier_completion["row_count"] == 1
            supplier_metric = next(
                item
                for item in analytics.json()["metrics"]
                if item["code"] == "supplier_change"
            )
            assert supplier_metric["value"] == 1
            assert supplier_metric["source_type"] == "supplier_module"
    finally:
        cleanup()


def test_supplier_excel_template_import_and_export() -> None:
    cleanup()
    try:
        with TestClient(app) as client:
            login(client)
            template = client.get(f"{settings.api_prefix}/suppliers/template")
            assert template.status_code == 200
            assert template.headers["cache-control"] == "no-store"
            template_book = load_workbook(BytesIO(template.content))
            template_sheet = template_book.active
            assert [cell.value for cell in template_sheet[1]] == [
                "序号",
                "供应商名称",
                "供应商联系人",
                "联系电话",
                "联系地址",
                "合作时间",
                "常用产品类型",
                "状态",
                "供应商备注",
                "变更说明",
            ]
            assert template_sheet["A2"].value == '=IF(B2="","",ROW()-1)'
            assert template_sheet.tables["SupplierImportTable"].ref == "A1:J500"
            assert len(template_sheet.data_validations.dataValidation) == 1

            workbook = template_book
            sheet = template_sheet
            sheet["B2"] = TEST_EXCEL_NAME
            sheet["C2"] = "Excel联系人"
            sheet["D2"] = "13700000000"
            sheet["E2"] = "Excel测试地址"
            sheet["F2"] = date(2032, 3, 8)
            sheet["G2"] = "包装材料,纸箱"
            sheet["B3"] = TEST_BATCH_EXCEL_NAME
            sheet["C3"] = "批量联系人"
            sheet["D3"] = "13600000000"
            sheet["E3"] = "批量测试地址"
            sheet["F3"] = date(2032, 3, 9)
            sheet["G3"] = "气泡膜,胶带"
            upload = BytesIO()
            workbook.save(upload)
            imported = client.post(
                f"{settings.api_prefix}/suppliers/import",
                params={"month": TEST_MONTH},
                files={
                    "file": (
                        "suppliers.xlsx",
                        upload.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            assert imported.status_code == 200
            assert imported.json()["created_count"] == 2
            assert imported.json()["total_count"] == 2

            exported = client.get(
                f"{settings.api_prefix}/suppliers/export",
                params={"keyword": TEST_EXCEL_NAME},
            )
            assert exported.status_code == 200
            export_book = load_workbook(BytesIO(exported.content), data_only=True)
            export_sheet = export_book.active
            assert export_sheet.max_column == 10
            assert export_sheet["B2"].value == TEST_EXCEL_NAME
            assert export_sheet["C2"].value == "Excel联系人"
            assert export_sheet["G2"].value == "包装材料,纸箱"
            assert export_sheet["H2"].value == "合作中"
            assert export_sheet.tables["SupplierExportTable"].ref == "A1:J2"

            listed = client.get(
                f"{settings.api_prefix}/suppliers",
                params={"month": TEST_MONTH, "keyword": TEST_EXCEL_NAME},
            )
            supplier_id = listed.json()["records"][0]["id"]
            manual_update = client.put(
                f"{settings.api_prefix}/suppliers/{supplier_id}",
                json={
                    "name": TEST_EXCEL_NAME,
                    "contact_name": "Excel联系人",
                    "contact_phone": "13700000000",
                    "address": "Excel测试地址",
                    "cooperation_start_date": "2032-03-08",
                    "product_types": "包装材料,纸箱",
                    "note": "仅在后台维护的备注",
                    "change_month": TEST_MONTH,
                    "change_note": "补充后台备注",
                },
            )
            assert manual_update.status_code == 200
            deactivated = client.patch(
                f"{settings.api_prefix}/suppliers/{supplier_id}/status",
                json={
                    "is_active": False,
                    "change_month": TEST_MONTH,
                    "change_note": "暂停合作",
                },
            )
            assert deactivated.status_code == 200

            sheet["C2"] = "更新联系人"
            updated_upload = BytesIO()
            workbook.save(updated_upload)
            updated = client.post(
                f"{settings.api_prefix}/suppliers/import",
                params={"month": TEST_MONTH},
                files={
                    "file": (
                        "suppliers_update.xlsx",
                        updated_upload.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            assert updated.status_code == 200
            assert updated.json()["updated_count"] == 1

            listed = client.get(
                f"{settings.api_prefix}/suppliers",
                params={"month": TEST_MONTH, "keyword": TEST_EXCEL_NAME},
            )
            assert listed.status_code == 200
            assert listed.json()["records"][0]["contact_name"] == "更新联系人"
            assert listed.json()["records"][0]["note"] == "仅在后台维护的备注"
            assert listed.json()["records"][0]["is_active"] is False
    finally:
        cleanup()
