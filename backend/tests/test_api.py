"""First-version API integration tests."""

from io import BytesIO
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, or_, select

from app.core.config import settings
from app.core.permissions import PERMISSION_CODES
from app.db.session import SessionLocal
from app.main import app
from app.models.reimbursement import Reimbursement, ReimbursementAttachment, ReimbursementEntity
from app.models.user import AuditLog, Role, RolePermission, User
from app.services import query_jobs
from app.services.invoice_ocr import parse_baidu_invoice


def login(client: TestClient) -> None:
    response = client.post(
        f"{settings.api_prefix}/auth/login",
        json={
            "username": settings.initial_admin_username,
            "password": settings.initial_admin_password,
        },
    )
    assert response.status_code == 200


def test_health_and_login_flow() -> None:
    with TestClient(app) as client:
        health = client.get(f"{settings.api_prefix}/health")
        assert health.status_code == 200
        assert health.json() == {"status": "ok"}

        unauthorized = client.get(f"{settings.api_prefix}/auth/me")
        assert unauthorized.status_code == 401

        login(client)

        current = client.get(f"{settings.api_prefix}/auth/me")
        assert current.status_code == 200


def test_migrated_modules_are_readable() -> None:
    with TestClient(app) as client:
        login(client)

        overview = client.get(f"{settings.api_prefix}/system/overview")
        assert overview.status_code == 200
        payload = overview.json()
        assert payload["database"] == "yibo_backoffice"
        assert {item["status"] for item in payload["modules"]} == {"ready"}
        assert payload["analytics"]["metric_definitions"] == 15
        assert payload["analytics"]["metric_values"] >= 63

        analytics = client.get(f"{settings.api_prefix}/analytics?month=2026-07")
        assert analytics.status_code == 200
        shipping = next(
            item for item in analytics.json()["metrics"] if item["code"] == "shipping_orders"
        )
        assert shipping["value"] > 0
        assert shipping["previous_value"] > 0

        history = client.get(f"{settings.api_prefix}/express/history")
        assert history.status_code == 200
        assert {item["month"] for item in history.json()} >= {
            "2026-05",
            "2026-06",
            "2026-07",
        }

        stats = client.get(f"{settings.api_prefix}/express/stats/2026-07")
        assert stats.status_code == 200
        assert stats.json()["total_orders"] > 0
        assert stats.json()["team_summary"][-1]["team"] == "合计"
        assert stats.json()["team_summary"][-1]["total_amount"] == pytest.approx(
            stats.json()["total_amount"]
        )

        unmatched = client.get(f"{settings.api_prefix}/express/unmatched/2026-07")
        assert unmatched.status_code == 200
        assert (
            unmatched.json()["matched"] + unmatched.json()["unmatched"] == unmatched.json()["total"]
        )

        preview = client.get(
            f"{settings.api_prefix}/express/preview/2026-07",
            params={"filter": "unmatched", "page": 1, "size": 10},
        )
        assert preview.status_code == 200
        assert preview.json()["filtered"] == unmatched.json()["unmatched"]
        assert all(row["所属团队"] == "未匹配" for row in preview.json()["rows"])

        rejected_upload = client.post(
            f"{settings.api_prefix}/express/upload",
            data={"month": "2026-07"},
            files={"file": ("unsupported.xls", b"legacy", "application/vnd.ms-excel")},
        )
        assert rejected_upload.status_code == 400

        anomaly_export = client.get(f"{settings.api_prefix}/express/anomalies/2026-07/download")
        assert anomaly_export.status_code == 200
        anomaly_book = load_workbook(BytesIO(anomaly_export.content), read_only=True)
        anomaly_headers = [cell.value for cell in anomaly_book.active[1]]
        assert anomaly_headers[-2:] == ["异常类型", "异常原因说明"]

        configs = client.get(f"{settings.api_prefix}/query-export/configs")
        assert configs.status_code == 200
        assert sum(len(items) for items in configs.json()["groups"].values()) == 4

        salary = client.get(f"{settings.api_prefix}/salary?month=2026-07")
        assert salary.status_code == 200
        assert salary.json()["records"] == []


@pytest.mark.parametrize(
    "sql",
    [
        "DELETE FROM orders",
        "UPDATE orders SET status = 1",
        "SELECT 1; DROP TABLE orders",
    ],
)
def test_query_export_rejects_write_sql(sql: str) -> None:
    with pytest.raises(ValueError):
        query_jobs.validate_read_query(sql)


def test_baidu_invoice_result_is_normalized_for_manual_confirmation() -> None:
    result = parse_baidu_invoice(
        {
            "log_id": 123456,
            "words_result": {
                "PurchaserName": "杭州测试供应链有限公司",
                "PurchaserRegisterNum": " 91330100 TEST000001 ",
                "AmountInFiguers": "￥1,234.56",
                "InvoiceCode": "033002400001",
                "InvoiceNum": "12345678",
                "InvoiceDate": "2026年08月29日",
            },
        }
    )
    assert result.entity_name == "杭州测试供应链有限公司"
    assert result.tax_number == "91330100TEST000001"
    assert float(result.amount or 0) == 1234.56
    assert result.invoice_number == "12345678"
    assert result.invoice_date is not None
    assert result.status == "success"


def test_reimbursement_full_workflow() -> None:
    claim_id: int | None = None
    entity_id: int | None = None
    with TestClient(app) as client:
        login(client)
        entity_name = f"接口测试主体_{uuid4().hex[:8]}"
        entity = client.post(
            f"{settings.api_prefix}/reimbursements/entities",
            json={
                "name": entity_name,
                "tax_number": "91330100TEST000001",
                "is_default": False,
                "is_active": True,
            },
        )
        assert entity.status_code == 201
        entity_id = entity.json()["id"]
        payload = {
            "applicant_name": "接口测试员工",
            "team": "发货组",
            "entity_name": entity_name,
            "tax_number": "91330100TEST000001",
            "note": "自动化测试完成后清理",
            "items": [
                {
                    "expense_date": "2026-08-27",
                    "category": "临时运费",
                    "amount": 12.5,
                    "related_number": "TEST-001",
                    "description": "测试明细",
                }
            ],
        }
        try:
            created = client.post(f"{settings.api_prefix}/reimbursements", json=payload)
            assert created.status_code == 201
            claim_id = created.json()["id"]
            assert created.json()["status"] == "draft"

            draft_export = client.get(
                f"{settings.api_prefix}/reimbursements/export/xlsx?ids={claim_id}"
            )
            assert draft_export.status_code == 200
            draft_book = load_workbook(BytesIO(draft_export.content), read_only=True)
            draft_sheet = draft_book["报销数据明细"]
            draft_headers = [cell.value for cell in draft_sheet[1]]
            assert draft_headers[:9] == [
                "报销单ID",
                "明细ID",
                "报销单号",
                "报销人",
                "所属组",
                "主体",
                "税号",
                "状态代码",
                "单据状态",
            ]
            assert draft_sheet.cell(2, 6).value == entity_name
            assert draft_sheet.cell(2, 7).value == "91330100TEST000001"
            assert draft_sheet.cell(2, 8).value == "draft"
            assert draft_sheet.cell(2, 9).value == "草稿"
            assert draft_sheet.cell(1, 22).value == "创建时间"
            assert draft_sheet.cell(2, 22).value is not None

            draft_marked = client.post(
                f"{settings.api_prefix}/reimbursements/export/mark", json={"ids": [claim_id]}
            )
            assert draft_marked.status_code == 200
            assert (
                client.get(f"{settings.api_prefix}/reimbursements/{claim_id}").json()["exported"]
                is True
            )

            attachment = client.post(
                f"{settings.api_prefix}/reimbursements/{claim_id}/attachments",
                files={"file": ("receipt.png", b"not-a-real-image", "image/png")},
            )
            assert attachment.status_code == 201
            assert (
                client.get(f"{settings.api_prefix}/reimbursements/{claim_id}").json()["exported"]
                is False
            )

            invoice_upload = client.post(
                f"{settings.api_prefix}/reimbursements/{claim_id}/attachments",
                data={"document_type": "invoice"},
                files={"file": ("invoice.png", b"not-a-real-invoice", "image/png")},
            )
            assert invoice_upload.status_code == 201
            invoice = invoice_upload.json()["invoice"]
            assert invoice["recognition_status"] == "unconfigured"
            corrected = client.put(
                f"{settings.api_prefix}/reimbursements/{claim_id}/invoices/{invoice['id']}",
                json={
                    "entity_name": entity_name,
                    "tax_number": "91330100TEST000001",
                    "amount": 12.5,
                },
            )
            assert corrected.status_code == 200
            assert corrected.json()["invoice"]["recognition_status"] == "confirmed"
            assert corrected.json()["invoice"]["manually_edited"] is True
            detail = client.get(f"{settings.api_prefix}/reimbursements/{claim_id}")
            assert detail.json()["invoice_count"] == 1
            assert detail.json()["invoice_amount"] == 12.5
            assert detail.json()["invoice_amount_difference"] == 0

            submitted = client.post(f"{settings.api_prefix}/reimbursements/{claim_id}/submit")
            assert submitted.status_code == 200
            assert submitted.json()["status"] == "pending_supervisor"

            approved = client.post(
                f"{settings.api_prefix}/reimbursements/{claim_id}/approve",
                json={"comment": "主管测试通过"},
            )
            assert approved.status_code == 200
            if approved.json()["status"] == "pending_finance":
                approved = client.post(
                    f"{settings.api_prefix}/reimbursements/{claim_id}/approve",
                    json={"comment": "财务测试通过"},
                )
                assert approved.status_code == 200
            assert approved.json()["status"] == "approved"

            exported = client.get(
                f"{settings.api_prefix}/reimbursements/export/xlsx?ids={claim_id}"
            )
            assert exported.status_code == 200
            assert exported.headers["content-type"].startswith(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            marked = client.post(
                f"{settings.api_prefix}/reimbursements/export/mark", json={"ids": [claim_id]}
            )
            assert marked.status_code == 200
            assert marked.json()["count"] == 1

            repeated_export = client.get(
                f"{settings.api_prefix}/reimbursements/export/xlsx?ids={claim_id}"
            )
            assert repeated_export.status_code == 200
            repeated_mark = client.post(
                f"{settings.api_prefix}/reimbursements/export/mark", json={"ids": [claim_id]}
            )
            assert repeated_mark.status_code == 200
            assert repeated_mark.json()["count"] == 1

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["费用日期", "费用类别", "金额", "关联单号", "费用说明"])
            sheet.append(["2026-08-27", "退件运费", 18.6, "TEST-002", "导入测试"])
            buffer = BytesIO()
            workbook.save(buffer)
            preview = client.post(
                f"{settings.api_prefix}/reimbursements/import/preview",
                files={
                    "file": (
                        "import.xlsx",
                        buffer.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            assert preview.status_code == 200
            assert preview.json()["count"] == 1
        finally:
            if claim_id is not None:
                with SessionLocal() as db:
                    attachments = db.scalars(
                        select(ReimbursementAttachment).where(
                            ReimbursementAttachment.reimbursement_id == claim_id
                        )
                    ).all()
                    root = settings.storage_path.resolve()
                    for attachment_row in attachments:
                        path = (settings.storage_path / attachment_row.relative_path).resolve()
                        if root in path.parents and path.is_file():
                            path.unlink()
                    db.execute(delete(Reimbursement).where(Reimbursement.id == claim_id))
                    if entity_id is not None:
                        db.execute(
                            delete(ReimbursementEntity).where(ReimbursementEntity.id == entity_id)
                        )
                    db.commit()


def test_finance_online_approval_and_admin_all_approval() -> None:
    claim_ids: list[int] = []
    finance_user_id: int | None = None
    finance_display_name = f"财务审批测试_{uuid4().hex[:8]}"

    with TestClient(app) as admin_client:
        login(admin_client)
        current = admin_client.get(f"{settings.api_prefix}/auth/me")
        assert current.status_code == 200
        admin_permissions = current.json()["user"]["permissions"]
        assert set(admin_permissions) == PERMISSION_CODES
        assert set(admin_permissions.values()) == {"all"}

        config = admin_client.get(f"{settings.api_prefix}/reimbursements/config")
        assert config.status_code == 200
        original_finance_setting = config.json()["finance_approval_enabled"]

        try:
            enabled = admin_client.put(
                f"{settings.api_prefix}/reimbursements/config",
                json={"finance_approval_enabled": True},
            )
            assert enabled.status_code == 200

            created_user = admin_client.post(
                f"{settings.api_prefix}/access/users",
                json={
                    "display_name": finance_display_name,
                    "team": "发货组",
                    "roles": ["finance"],
                },
            )
            assert created_user.status_code == 201
            finance_user_id = created_user.json()["user"]["id"]
            finance_password = created_user.json()["temporary_password"]

            def create_pending_finance(note: str) -> int:
                payload = {
                    "applicant_name": "管理员流程测试",
                    "team": "发货组",
                    "entity_name": "财务审批测试主体",
                    "tax_number": "91330100FINANCE001",
                    "note": note,
                    "items": [
                        {
                            "expense_date": "2026-08-28",
                            "category": "临时运费",
                            "amount": 8.28,
                            "related_number": "ONLINE-FINANCE",
                            "description": "在线审批权限测试",
                        }
                    ],
                }
                created = admin_client.post(
                    f"{settings.api_prefix}/reimbursements", json=payload
                )
                assert created.status_code == 201
                claim_id = created.json()["id"]
                claim_ids.append(claim_id)

                submitted = admin_client.post(
                    f"{settings.api_prefix}/reimbursements/{claim_id}/submit"
                )
                assert submitted.status_code == 200
                supervisor_approved = admin_client.post(
                    f"{settings.api_prefix}/reimbursements/{claim_id}/approve",
                    json={"comment": "管理员执行主管审批"},
                )
                assert supervisor_approved.status_code == 200
                assert supervisor_approved.json()["status"] == "pending_finance"
                return claim_id

            finance_claim_id = create_pending_finance("财务角色在线审批")

            exported = admin_client.get(
                f"{settings.api_prefix}/reimbursements/export/xlsx",
                params={"ids": str(finance_claim_id)},
            )
            assert exported.status_code == 200
            unchanged = admin_client.get(
                f"{settings.api_prefix}/reimbursements/{finance_claim_id}"
            )
            assert unchanged.json()["status"] == "pending_finance"

            with TestClient(app) as finance_client:
                signed_in = finance_client.post(
                    f"{settings.api_prefix}/auth/login",
                    json={
                        "username": finance_display_name,
                        "password": finance_password,
                    },
                )
                assert signed_in.status_code == 200
                permissions = signed_in.json()["user"]["permissions"]
                assert permissions["reimbursement.approve_finance"] == "all"
                assert "reimbursement.approve_supervisor" not in permissions

                finance_detail = finance_client.get(
                    f"{settings.api_prefix}/reimbursements/{finance_claim_id}"
                )
                assert finance_detail.status_code == 200
                assert finance_detail.json()["can_approve"] is True
                finance_approved = finance_client.post(
                    f"{settings.api_prefix}/reimbursements/{finance_claim_id}/approve",
                    json={"comment": "财务角色在线审批通过"},
                )
                assert finance_approved.status_code == 200
                assert finance_approved.json()["status"] == "approved"
                assert finance_approved.json()["approval_records"][-1]["action"] == (
                    "finance_approve"
                )

            admin_claim_id = create_pending_finance("管理员完成全部审批")
            admin_approved = admin_client.post(
                f"{settings.api_prefix}/reimbursements/{admin_claim_id}/approve",
                json={"comment": "管理员执行财务审批"},
            )
            assert admin_approved.status_code == 200
            assert admin_approved.json()["status"] == "approved"
            assert admin_approved.json()["approval_records"][-1]["action"] == (
                "finance_approve"
            )
        finally:
            restored = admin_client.put(
                f"{settings.api_prefix}/reimbursements/config",
                json={"finance_approval_enabled": original_finance_setting},
            )
            assert restored.status_code == 200
            with SessionLocal() as db:
                if claim_ids:
                    db.execute(delete(Reimbursement).where(Reimbursement.id.in_(claim_ids)))
                if finance_user_id is not None:
                    db.execute(
                        delete(AuditLog).where(
                            or_(
                                AuditLog.user_id == finance_user_id,
                                AuditLog.resource == f"user:{finance_user_id}",
                            )
                        )
                    )
                    db.execute(delete(User).where(User.id == finance_user_id))
                db.commit()


def test_reimbursement_batch_import_and_template() -> None:
    claim_ids: list[int] = []
    with TestClient(app) as client:
        login(client)
        current = client.get(f"{settings.api_prefix}/auth/me")
        applicant_name = current.json()["user"]["display_name"]

        template = client.get(f"{settings.api_prefix}/reimbursements/batch/template/xlsx")
        assert template.status_code == 200
        assert template.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "批量导入"
        sheet.append(["报销批量导入模板"])
        sheet.append(["填写说明"])
        sheet.append([])
        sheet.append(
            [
                "报销分组*",
                "报销人*",
                "所属组*",
                "主体*",
                "税号*",
                "费用日期*",
                "费用类别*",
                "金额*",
                "关联单号",
                "费用说明",
                "整单备注",
            ]
        )
        sheet.append(
            [
                "TEST001",
                applicant_name,
                "发货组",
                "批量测试主体",
                "91330100BATCH0001",
                "2026-08-28",
                "临时运费",
                12.5,
                "BATCH-001",
                "第一条",
                "批量导入测试",
            ]
        )
        sheet.append(
            [
                "TEST001",
                applicant_name,
                "发货组",
                "批量测试主体",
                "91330100BATCH0001",
                "2026-08-28",
                "包材临时采购",
                8,
                "BATCH-002",
                "第二条",
                "批量导入测试",
            ]
        )
        sheet.append(
            [
                "TEST002",
                applicant_name,
                "退货组",
                "批量测试主体",
                "91330100BATCH0001",
                "2026-08-28",
                "退件运费",
                6,
                "BATCH-003",
                "第三条",
                "",
            ]
        )
        buffer = BytesIO()
        workbook.save(buffer)
        contents = buffer.getvalue()

        try:
            preview = client.post(
                f"{settings.api_prefix}/reimbursements/batch/preview",
                files={
                    "file": (
                        "batch.xlsx",
                        contents,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            assert preview.status_code == 200
            assert preview.json()["can_import"] is True
            assert preview.json()["claim_count"] == 2
            assert preview.json()["item_count"] == 3
            assert preview.json()["total_amount"] == 26.5

            imported = client.post(
                f"{settings.api_prefix}/reimbursements/batch/import",
                data={"submit": "false"},
                files={
                    "file": (
                        "batch.xlsx",
                        contents,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            assert imported.status_code == 201
            assert imported.json()["claim_count"] == 2
            assert imported.json()["item_count"] == 3
            assert imported.json()["status"] == "draft"
            claim_ids = [item["id"] for item in imported.json()["records"]]

            draft_export = client.get(
                f"{settings.api_prefix}/reimbursements/export/xlsx?ids={claim_ids[0]}"
            )
            assert draft_export.status_code == 200
            marked = client.post(
                f"{settings.api_prefix}/reimbursements/export/mark",
                json={"ids": [claim_ids[0]]},
            )
            assert marked.status_code == 200
            exported_rows = client.get(
                f"{settings.api_prefix}/reimbursements?view=exported"
            ).json()["records"]
            pending_rows = client.get(
                f"{settings.api_prefix}/reimbursements?view=pending_export"
            ).json()["records"]
            assert claim_ids[0] in {row["id"] for row in exported_rows}
            assert claim_ids[1] in {row["id"] for row in pending_rows}
        finally:
            if claim_ids:
                with SessionLocal() as db:
                    db.execute(delete(Reimbursement).where(Reimbursement.id.in_(claim_ids)))
                    db.commit()


def test_account_role_and_permission_lifecycle() -> None:
    role_code: str | None = None
    user_id: int | None = None

    try:
        with TestClient(app) as admin_client:
            login(admin_client)
            overview = admin_client.get(f"{settings.api_prefix}/access/overview")
            assert overview.status_code == 200
            assert overview.json()["summary"]["administrators"] >= 1

            permissions = admin_client.get(f"{settings.api_prefix}/access/permissions")
            assert permissions.status_code == 200
            assert {item["code"] for item in permissions.json()} >= {
                "dashboard.view",
                "accounts.manage",
                "reimbursement.view",
            }

            roles = admin_client.get(f"{settings.api_prefix}/access/roles")
            assert roles.status_code == 200
            team_leader = next(item for item in roles.json() if item["code"] == "team_leader")
            assert team_leader["name"] == "组长"
            assert team_leader["is_system"] is True
            assert team_leader["permissions"] == {
                "dashboard.view": "all",
                "reimbursement.view": "team",
                "reimbursement.create": "self",
                "reimbursement.approve_supervisor": "team",
            }

            role = admin_client.post(
                f"{settings.api_prefix}/access/roles",
                json={
                    "name": "接口测试角色",
                    "description": "自动化测试后清理",
                    "permissions": {
                        "dashboard.view": "all",
                        "accounts.view": "all",
                        "accounts.manage": "all",
                        "reimbursement.view": "self",
                        "reimbursement.create": "self",
                    },
                },
            )
            assert role.status_code == 201
            role_code = role.json()["code"]

            created = admin_client.post(
                f"{settings.api_prefix}/access/users",
                json={
                    "display_name": "权限测试员工",
                    "team": "发货组",
                    "roles": [role_code, "finance"],
                },
            )
            assert created.status_code == 201
            assert set(created.json()["user"]["roles"]) == {role_code, "finance"}
            assert created.json()["user"]["latest_password"] == "423766"
            user_id = created.json()["user"]["id"]
            first_password = created.json()["temporary_password"]
            assert first_password == "423766"

            duplicate_name = admin_client.post(
                f"{settings.api_prefix}/access/users",
                json={
                    "display_name": "权限测试员工",
                    "team": "退货组",
                    "roles": ["employee"],
                },
            )
            assert duplicate_name.status_code == 409

            with TestClient(app) as user_client:
                signed_in = user_client.post(
                    f"{settings.api_prefix}/auth/login",
                    json={"username": "权限测试员工", "password": first_password},
                )
                assert signed_in.status_code == 200
                assert signed_in.json()["user"]["must_change_password"] is False
                assert set(signed_in.json()["user"]["roles"]) == {role_code, "finance"}
                assert signed_in.json()["user"]["permissions"]["reimbursement.view"] == "all"

                first_session = user_client.get(f"{settings.api_prefix}/system/overview")
                assert first_session.status_code == 403

                signed_out = user_client.post(f"{settings.api_prefix}/auth/logout")
                assert signed_out.status_code == 204
                signed_in_again = user_client.post(
                    f"{settings.api_prefix}/auth/login",
                    json={"username": "权限测试员工", "password": first_password},
                )
                assert signed_in_again.status_code == 200
                assert signed_in_again.json()["user"]["must_change_password"] is True
                skipped = user_client.get(f"{settings.api_prefix}/system/overview")
                assert skipped.status_code == 403

                changed = user_client.post(
                    f"{settings.api_prefix}/auth/change-password",
                    json={
                        "current_password": first_password,
                        "new_password": "654321",
                    },
                )
                assert changed.status_code == 200
                assert changed.json()["user"]["must_change_password"] is False

                managed_users = admin_client.get(f"{settings.api_prefix}/access/users")
                assert managed_users.status_code == 200
                managed_user = next(item for item in managed_users.json() if item["id"] == user_id)
                assert managed_user["latest_password"] == "654321"

                dashboard = user_client.get(f"{settings.api_prefix}/system/overview")
                assert dashboard.status_code == 403
                salary = user_client.get(f"{settings.api_prefix}/salary?month=2026-07")
                assert salary.status_code == 200
                access = user_client.get(f"{settings.api_prefix}/access/overview")
                assert access.status_code == 403

                disabled = admin_client.put(
                    f"{settings.api_prefix}/access/users/{user_id}",
                    json={
                        "display_name": "权限测试员工",
                        "team": "发货组",
                        "roles": [role_code, "finance"],
                        "is_active": False,
                    },
                )
                assert disabled.status_code == 200
                assert user_client.get(f"{settings.api_prefix}/auth/me").status_code == 401
    finally:
        with SessionLocal() as db:
            filters = []
            if user_id is not None:
                filters.extend(
                    [AuditLog.user_id == user_id, AuditLog.resource == f"user:{user_id}"]
                )
            if role_code:
                filters.append(AuditLog.resource == f"role:{role_code}")
            if filters:
                db.execute(delete(AuditLog).where(or_(*filters)))
            if user_id is not None:
                db.execute(delete(User).where(User.id == user_id))
            if role_code:
                db.execute(delete(RolePermission).where(RolePermission.role_code == role_code))
                db.execute(delete(Role).where(Role.code == role_code))
            db.commit()
