"""Operating-analysis detail import tests."""

from datetime import date, datetime
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, select

from app.api.v1 import analytics as analytics_api
from app.core.config import settings
from app.db.session import SessionLocal
from app.main import app
from app.models.analytics import (
    AnalyticsDetailRow,
    AnalyticsImportBatch,
    MetricDefinition,
    MonthlyMetric,
    MonthlyReview,
)
from app.models.user import AuditLog
from app.services import analytics_sources


def workbook_bytes(values: tuple[int, int]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "1-发货单量"
    sheet.append(["月份", "团队名称", "发货单量", "备注"])
    sheet.append(["2031-01", "测试一组", values[0], "人工上传"])
    sheet.append(["2031-01", "测试二组", values[1], "人工上传"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def staffing_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "5-人员调整"
    sheet.append(
        [
            "月份",
            "小组",
            "正式工人数",
            "最优配置",
            "配置偏差",
            "偏差比例",
            "人均月产出",
            "最低人均产出",
            "效率损失",
            "效率损失占比",
            "人均月产出变化",
            "人均月产出环比",
            "分析",
        ]
    )
    sheet.append(
        [
            "2099-12",
            "测试小组",
            10.5,
            10,
            999,
            0.99,
            1500,
            1200,
            999,
            0.99,
            999,
            0.99,
            "人员配置略高",
        ]
    )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def login(client: TestClient) -> None:
    response = client.post(
        f"{settings.api_prefix}/auth/login",
        json={
            "username": settings.initial_admin_username,
            "password": settings.initial_admin_password,
        },
    )
    assert response.status_code == 200


def cleanup_test_month() -> None:
    selected = date(2031, 1, 1)
    with SessionLocal() as db:
        batch_ids = list(
            db.scalars(
                select(AnalyticsImportBatch.id).where(
                    AnalyticsImportBatch.dataset_type == "shipping_orders",
                    AnalyticsImportBatch.month == selected,
                )
            )
        )
        if batch_ids:
            db.execute(
                delete(AnalyticsDetailRow).where(AnalyticsDetailRow.batch_id.in_(batch_ids))
            )
            db.execute(
                delete(AnalyticsImportBatch).where(AnalyticsImportBatch.id.in_(batch_ids))
            )
        metric_id = db.scalar(
            select(MetricDefinition.id).where(MetricDefinition.code == "shipping_orders")
        )
        if metric_id is not None:
            db.execute(
                delete(MonthlyMetric).where(
                    MonthlyMetric.metric_id == metric_id,
                    MonthlyMetric.month == selected,
                )
            )
        db.execute(
            delete(AuditLog).where(
                AuditLog.resource.in_(
                    ["analytics:shipping_orders:2031-01", "analytics:2031-01"]
                )
            )
        )
        db.execute(delete(MonthlyReview).where(MonthlyReview.month == selected))
        db.commit()


def cleanup_staffing_test_month() -> None:
    selected = date(2031, 1, 1)
    with SessionLocal() as db:
        batch_ids = list(
            db.scalars(
                select(AnalyticsImportBatch.id).where(
                    AnalyticsImportBatch.dataset_type == "staffing",
                    AnalyticsImportBatch.month == selected,
                )
            )
        )
        if batch_ids:
            db.execute(
                delete(AnalyticsDetailRow).where(AnalyticsDetailRow.batch_id.in_(batch_ids))
            )
            db.execute(
                delete(AnalyticsImportBatch).where(AnalyticsImportBatch.id.in_(batch_ids))
            )
        metric_id = db.scalar(
            select(MetricDefinition.id).where(MetricDefinition.code == "staff_adjustment")
        )
        if metric_id is not None:
            db.execute(
                delete(MonthlyMetric).where(
                    MonthlyMetric.metric_id == metric_id,
                    MonthlyMetric.month == selected,
                )
            )
        db.execute(
            delete(AuditLog).where(
                AuditLog.resource.like("analytics:staffing:2031-01%")
            )
        )
        db.commit()


def cleanup_return_test_month() -> None:
    selected = date(2031, 2, 1)
    previous = date(2031, 1, 1)
    with SessionLocal() as db:
        batch_ids = list(
            db.scalars(
                select(AnalyticsImportBatch.id).where(
                    AnalyticsImportBatch.dataset_type == "return_items",
                    AnalyticsImportBatch.month == selected,
                )
            )
        )
        if batch_ids:
            db.execute(
                delete(AnalyticsDetailRow).where(AnalyticsDetailRow.batch_id.in_(batch_ids))
            )
            db.execute(
                delete(AnalyticsImportBatch).where(AnalyticsImportBatch.id.in_(batch_ids))
            )
        metric_id = db.scalar(
            select(MetricDefinition.id).where(MetricDefinition.code == "return_items")
        )
        if metric_id is not None:
            db.execute(
                delete(MonthlyMetric).where(
                    MonthlyMetric.metric_id == metric_id,
                    MonthlyMetric.month.in_([previous, selected]),
                )
            )
        db.execute(
            delete(AuditLog).where(AuditLog.resource == "analytics:return_items:2031-02")
        )
        db.execute(delete(MonthlyReview).where(MonthlyReview.month == selected))
        db.commit()


def test_staffing_export_includes_headers_and_current_month_rows() -> None:
    cleanup_staffing_test_month()
    try:
        with TestClient(app) as client:
            login(client)

            empty_export = client.get(
                f"{settings.api_prefix}/analytics/details/staffing/export",
                params={"month": "2031-01"},
            )
            assert empty_export.status_code == 200
            empty_book = load_workbook(BytesIO(empty_export.content), data_only=True)
            empty_sheet = empty_book["人员调整"]
            assert empty_sheet.max_row == 1
            assert [cell.value for cell in empty_sheet[1]][:3] == [
                "小组",
                "正式工人数",
                "最优配置",
            ]

            template = client.get(
                f"{settings.api_prefix}/analytics/details/staffing/template"
            )
            assert template.status_code == 200
            template_book = load_workbook(BytesIO(template.content), data_only=False)
            template_sheet = template_book["人员调整"]
            assert [cell.value for cell in template_sheet[1]] == [
                "小组",
                "正式工人数",
                "最优配置",
                "人均月产出",
                "最优人均产出",
                "综合分析",
            ]

            empty_details = client.get(
                f"{settings.api_prefix}/analytics/details/staffing",
                params={"month": "2031-01"},
            )
            assert empty_details.status_code == 200
            empty_payload = empty_details.json()
            assert empty_payload["is_template"] is True
            assert empty_payload["columns"] == [
                "小组",
                "正式工人数",
                "最优配置",
                "配置偏差",
                "偏差比例",
                "人均月产出",
                "最优人均产出",
                "效率差额",
                "效率差额占比",
                "人均月产出净变化",
                "人均月产出环比",
                "综合分析",
            ]
            assert {row["values"]["小组"] for row in empty_payload["rows"]} == {
                "发货组",
                "售后组",
            }
            assert all(row["id"] < 0 for row in empty_payload["rows"])
            assert all(
                row["values"]["正式工人数"] is None
                for row in empty_payload["rows"]
            )
            with SessionLocal() as db:
                assert db.scalar(
                    select(AnalyticsImportBatch).where(
                        AnalyticsImportBatch.dataset_type == "staffing",
                        AnalyticsImportBatch.month == date(2031, 1, 1),
                    )
                ) is None

            virtual_row = next(
                row for row in empty_payload["rows"] if row["values"]["小组"] == "发货组"
            )
            initialized = client.patch(
                f"{settings.api_prefix}/analytics/details/staffing/rows/"
                f"{virtual_row['id']}/inputs",
                json={
                    "month": "2031-01",
                    "team_name": "发货组",
                    "regular_staff": 9,
                    "optimal_staff": 8,
                    "monthly_output": 10000,
                    "optimal_monthly_output": 9000,
                },
            )
            assert initialized.status_code == 200
            assert initialized.json()["row_id"] > 0
            assert initialized.json()["regular_total"] == 9
            initialized_details = client.get(
                f"{settings.api_prefix}/analytics/details/staffing",
                params={"month": "2031-01"},
            ).json()
            assert initialized_details["is_template"] is False
            assert initialized_details["total"] == 2
            assert initialized_details["batches"][0]["mode"] == "manual"
            initialized_rows = {
                row["values"]["小组"]: row["values"]
                for row in initialized_details["rows"]
            }
            assert initialized_rows["发货组"]["配置偏差"] == 1
            assert initialized_rows["售后组"]["正式工人数"] is None

            imported = client.post(
                f"{settings.api_prefix}/analytics/details/staffing/import",
                data={"month": "2031-01", "mode": "replace"},
                files={
                    "file": (
                        "人员调整.xlsx",
                        staffing_workbook_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            assert imported.status_code == 200
            assert imported.json()["warnings"][-1] == (
                "已按页面月份 2031-01 保存，派生指标由系统重新计算。"
            )

            details = client.get(
                f"{settings.api_prefix}/analytics/details/staffing",
                params={"month": "2031-01"},
            )
            assert details.status_code == 200
            staffing_row = next(
                row
                for row in details.json()["rows"]
                if row["values"]["小组"] == "测试小组"
            )
            analysis = (
                "1、配置分析。人员配置已调整至合理区间\n"
                "2、人均月产出分析。人均产出保持稳定\n"
                "3、人均月产出环比分析。环比小幅提升"
            )
            updated = client.patch(
                f"{settings.api_prefix}/analytics/details/staffing/rows/"
                f"{staffing_row['id']}/analysis",
                json={"month": "2031-01", "analysis": analysis},
            )
            assert updated.status_code == 200
            assert updated.json()["analysis"] == analysis
            refreshed = client.get(
                f"{settings.api_prefix}/analytics/details/staffing",
                params={"month": "2031-01"},
            )
            refreshed_staffing_row = next(
                row
                for row in refreshed.json()["rows"]
                if row["values"]["小组"] == "测试小组"
            )
            assert refreshed_staffing_row["values"]["综合分析"] == analysis

            exported = client.get(
                f"{settings.api_prefix}/analytics/details/staffing/export",
                params={"month": "2031-01"},
            )
            assert exported.status_code == 200
            assert "2031-01%20%E4%BA%BA%E5%91%98%E8%B0%83%E6%95%B4" in exported.headers[
                "content-disposition"
            ]
            export_book = load_workbook(BytesIO(exported.content), data_only=True)
            export_sheet = export_book["人员调整"]
            exported_staffing_row = next(
                [cell.value for cell in row]
                for row in export_sheet.iter_rows(min_row=2)
                if row[0].value == "测试小组"
            )
            assert exported_staffing_row == [
                "测试小组",
                10.5,
                10,
                0.5,
                0.047619,
                1500,
                1200,
                300,
                0.25,
                None,
                None,
                analysis,
            ]

            inputs_updated = client.patch(
                f"{settings.api_prefix}/analytics/details/staffing/rows/"
                f"{staffing_row['id']}/inputs",
                json={
                    "month": "2031-01",
                    "regular_staff": 11.5,
                    "optimal_staff": 10,
                    "monthly_output": 1600,
                    "optimal_monthly_output": 1250,
                },
            )
            assert inputs_updated.status_code == 200
            values = inputs_updated.json()["values"]
            assert values["正式工人数"] == 11.5
            assert values["配置偏差"] == 1.5
            assert values["偏差比例"] == 0.130435
            assert values["效率差额"] == 350
            assert values["效率差额占比"] == 0.28
            assert values["人均月产出净变化"] is None
            assert values["人均月产出环比"] is None
            assert values["综合分析"] == analysis
            assert inputs_updated.json()["regular_total"] == 20.5
            with SessionLocal() as db:
                staffing_metric = db.scalar(
                    select(MonthlyMetric)
                    .join(MetricDefinition)
                    .where(
                        MonthlyMetric.month == date(2031, 1, 1),
                        MetricDefinition.code == "staff_adjustment",
                    )
                )
                assert staffing_metric is not None
                assert float(staffing_metric.value) == 20.5
                assert staffing_metric.source_type == "manual"

            second_import = client.post(
                f"{settings.api_prefix}/analytics/details/staffing/import",
                data={"month": "2031-01", "mode": "append"},
                files={
                    "file": (
                        "人员调整-修订.xlsx",
                        staffing_workbook_bytes(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            assert second_import.status_code == 200
            with SessionLocal() as db:
                batches = db.scalars(
                    select(AnalyticsImportBatch).where(
                        AnalyticsImportBatch.dataset_type == "staffing",
                        AnalyticsImportBatch.month == date(2031, 1, 1),
                    )
                ).all()
                assert len(batches) == 3
                assert sum(1 for batch in batches if batch.active) == 1
    finally:
        cleanup_staffing_test_month()


def test_analytics_detail_preview_import_and_replace() -> None:
    cleanup_test_month()
    try:
        with TestClient(app) as client:
            login(client)
            first_file = workbook_bytes((10, 20))
            preview = client.post(
                f"{settings.api_prefix}/analytics/details/shipping_orders/preview",
                files={"file": ("经营分析.xlsx", first_file, "application/vnd.ms-excel")},
            )
            assert preview.status_code == 200
            assert preview.json()["row_count"] == 2
            assert preview.json()["summary"]["shipping_orders"] == 30

            imported = client.post(
                f"{settings.api_prefix}/analytics/details/shipping_orders/import",
                data={"month": "2031-01", "mode": "replace"},
                files={"file": ("经营分析.xlsx", first_file, "application/vnd.ms-excel")},
            )
            assert imported.status_code == 200
            assert imported.json()["updated_metrics"][0]["value"] == 30

            replacement = workbook_bytes((2, 3))
            match_preview = client.post(
                f"{settings.api_prefix}/analytics/details/shipping_orders/preview",
                params={"month": "2031-01"},
                files={"file": ("替换.xlsx", replacement, "application/vnd.ms-excel")},
            )
            assert match_preview.status_code == 200
            assert match_preview.json()["match_result"]["matched_count"] == 2
            assert match_preview.json()["match_result"]["unmatched_count"] == 0
            assert match_preview.json()["summary"]["shipping_orders"] == 5

            replaced = client.post(
                f"{settings.api_prefix}/analytics/details/shipping_orders/import",
                data={"month": "2031-01", "mode": "replace"},
                files={"file": ("替换.xlsx", replacement, "application/vnd.ms-excel")},
            )
            assert replaced.status_code == 200
            assert replaced.json()["updated_metrics"][0]["value"] == 5

            details = client.get(
                f"{settings.api_prefix}/analytics/details/shipping_orders",
                params={"month": "2031-01"},
            )
            assert details.status_code == 200
            assert details.json()["total"] == 2
            assert details.json()["summary"]["shipping_orders"] == 5
            assert details.json()["batches"][0]["original_name"] == "Excel 匹配 · 替换.xlsx"
            assert details.json()["batches"][0]["mode"] == "match"

            analytics = client.get(
                f"{settings.api_prefix}/analytics", params={"month": "2031-01"}
            )
            assert analytics.status_code == 200
            shipping = next(
                item for item in analytics.json()["metrics"] if item["code"] == "shipping_orders"
            )
            assert shipping["value"] == 5
            assert shipping["source_type"] == "excel"
            shipping_completion = next(
                item
                for item in analytics.json()["completion"]["items"]
                if item["code"] == "shipping_orders"
            )
            assert shipping_completion["state"] == "uploaded"

            template = client.get(
                f"{settings.api_prefix}/analytics/details/shipping_orders/template",
                params={"month": "2031-01"},
            )
            assert template.status_code == 200
            assert template.headers["content-type"].startswith(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            template_book = load_workbook(BytesIO(template.content), data_only=False)
            template_sheet = template_book["发货单量"]
            assert [cell.value for cell in template_sheet[1]] == [
                "序号",
                "团队名称",
                "发货单量",
                "数据发货占比",
                "备注",
            ]
            assert template_sheet["B2"].value == "测试一组"
            assert template_sheet["C2"].value == 2
            assert template_sheet["D2"].value == "=IFERROR(C2/SUM($C$2:$C$3),0)"
            assert template_sheet["D2"].number_format == "0.00%"
    finally:
        cleanup_test_month()


def test_lightweight_monthly_review_and_archive_lock() -> None:
    cleanup_test_month()
    try:
        with TestClient(app) as client:
            login(client)
            initial = client.get(
                f"{settings.api_prefix}/analytics", params={"month": "2031-01"}
            )
            assert initial.status_code == 200
            shipping = next(
                item for item in initial.json()["metrics"] if item["code"] == "shipping_orders"
            )
            saved = client.put(
                f"{settings.api_prefix}/analytics",
                json={
                    "month": "2031-01",
                    "metrics": [
                        {"metric_id": shipping["id"], "value": 42, "note": "测试说明"}
                    ],
                    "summary": "整体平稳",
                    "highlights": "发货效率提升",
                    "issues": "人工登记仍较多",
                    "risks": "数据口径待统一",
                    "next_plan": "完成模板培训",
                },
            )
            assert saved.status_code == 200

            analytics = client.get(
                f"{settings.api_prefix}/analytics", params={"month": "2031-01"}
            ).json()
            assert analytics["review"]["highlights"] == "发货效率提升"
            assert analytics["review"]["next_plan"] == "完成模板培训"
            assert analytics["latest_activity"]["updated_by_name"] == "系统管理员"
            shipping = next(
                item for item in analytics["metrics"] if item["code"] == "shipping_orders"
            )
            assert shipping["source_type"] == "manual"
            shipping_completion = next(
                item
                for item in analytics["completion"]["items"]
                if item["code"] == "shipping_orders"
            )
            assert shipping_completion["state"] == "summary_only"

            completed = client.put(
                f"{settings.api_prefix}/analytics/status",
                json={"month": "2031-01", "status": "completed"},
            )
            assert completed.status_code == 200
            archived = client.put(
                f"{settings.api_prefix}/analytics/status",
                json={"month": "2031-01", "status": "archived"},
            )
            assert archived.status_code == 200

            locked = client.put(
                f"{settings.api_prefix}/analytics",
                json={
                    "month": "2031-01",
                    "metrics": [],
                    "summary": "不应保存",
                    "highlights": "",
                    "issues": "",
                    "risks": "",
                    "next_plan": "",
                },
            )
            assert locked.status_code == 409

            reopened = client.put(
                f"{settings.api_prefix}/analytics/status",
                json={"month": "2031-01", "status": "draft"},
            )
            assert reopened.status_code == 200
            assert reopened.json()["status"] == "draft"
    finally:
        cleanup_test_month()


def test_shipping_system_preview_and_sync(monkeypatch) -> None:
    cleanup_test_month()
    monkeypatch.setattr(
        analytics_api,
        "current_business_month",
        lambda: date(2031, 3, 1),
    )

    def fake_shipping_source(selected: date) -> dict[str, object]:
        assert selected == date(2031, 1, 1)
        return {
            "source_name": "yibo.pre_matched_order（团队名称：yibo.team_info）",
            "month_start": "2031-01-01",
            "month_end": "2031-02-01",
            "columns": ["团队名称", "发货单量", "数据发货占比", "备注"],
            "rows": [
                {
                    "团队名称": "测试一组",
                    "发货单量": 10,
                    "数据发货占比": 33.33,
                    "备注": "",
                },
                {
                    "团队名称": "测试二组",
                    "发货单量": 20,
                    "数据发货占比": 66.67,
                    "备注": "",
                },
            ],
            "row_count": 2,
            "total": 30,
            "conditions": [
                "action_status IN (10, 20, 22, 28, 29, 30)",
                "part_logistics <> 2",
                "warehouse_id IN (1, -1)",
                "COUNT(DISTINCT waybill_id)",
                "团队名称取自 team_info.name",
            ],
        }

    monkeypatch.setattr(analytics_api, "fetch_shipping_orders", fake_shipping_source)
    try:
        with TestClient(app) as client:
            login(client)
            preview = client.get(
                f"{settings.api_prefix}/analytics/details/shipping_orders/system-preview",
                params={"month": "2031-01"},
            )
            assert preview.status_code == 200
            assert preview.json()["month_start"] == "2031-01-01"
            assert preview.json()["month_end"] == "2031-02-01"
            assert preview.json()["row_count"] == 2
            assert preview.json()["total"] == 30
            assert preview.json()["snapshot"]["state"] == "review"
            assert preview.json()["snapshot"]["can_system_sync"] is True

            synced = client.post(
                f"{settings.api_prefix}/analytics/details/shipping_orders/system-sync",
                params={"month": "2031-01"},
            )
            assert synced.status_code == 200
            assert synced.json()["total"] == 30

            details = client.get(
                f"{settings.api_prefix}/analytics/details/shipping_orders",
                params={"month": "2031-01"},
            ).json()
            assert details["total"] == 2
            assert details["summary"]["shipping_orders"] == 30
            assert details["snapshot"]["state"] == "review"
            assert details["snapshot"]["version_count"] == 1
            assert details["batches"][0]["mode"] == "system"
            assert details["columns"] == [
                "团队名称",
                "发货单量",
                "数据发货占比",
                "备注",
            ]
            assert details["rows"][0]["values"] == {
                "团队名称": "测试一组",
                "发货单量": 10,
                "数据发货占比": 33.33,
                "备注": "",
            }

            searched = client.get(
                f"{settings.api_prefix}/analytics/details/shipping_orders",
                params={"month": "2031-01", "search": "二"},
            ).json()
            assert searched["total"] == 1
            assert searched["rows"][0]["values"]["团队名称"] == "测试二组"

            sorted_details = client.get(
                f"{settings.api_prefix}/analytics/details/shipping_orders",
                params={"month": "2031-01", "sort_order": "desc"},
            ).json()
            assert sorted_details["rows"][0]["values"]["发货单量"] == 20

            first_row_id = details["rows"][0]["id"]
            exported = client.post(
                f"{settings.api_prefix}/analytics/details/shipping_orders/export",
                json={
                    "month": "2031-01",
                    "scope": "selected",
                    "row_ids": [first_row_id],
                    "columns": ["团队名称", "发货单量", "数据发货占比"],
                    "search": "",
                    "sort_order": "",
                },
            )
            assert exported.status_code == 200
            export_book = load_workbook(BytesIO(exported.content), data_only=True)
            export_sheet = export_book["发货数据"]
            assert [cell.value for cell in export_sheet[1]] == [
                "序号",
                "团队名称",
                "发货单量",
                "数据发货占比",
            ]
            assert [cell.value for cell in export_sheet[2]][:3] == [1, "测试一组", 10]
            assert round(export_sheet["D2"].value, 4) == 0.3333

            updated_remark = client.patch(
                f"{settings.api_prefix}/analytics/details/shipping_orders/rows/"
                f"{first_row_id}/remark",
                json={"month": "2031-01", "remark": "重点客户，持续关注"},
            )
            assert updated_remark.status_code == 200
            assert updated_remark.json()["remark"] == "重点客户，持续关注"

            synced_again = client.post(
                f"{settings.api_prefix}/analytics/details/shipping_orders/system-sync",
                params={"month": "2031-01"},
            )
            assert synced_again.status_code == 200
            refreshed_details = client.get(
                f"{settings.api_prefix}/analytics/details/shipping_orders",
                params={"month": "2031-01"},
            ).json()
            assert refreshed_details["rows"][0]["values"]["备注"] == "重点客户，持续关注"

            analytics = client.get(
                f"{settings.api_prefix}/analytics", params={"month": "2031-01"}
            ).json()
            shipping = next(
                item for item in analytics["metrics"] if item["code"] == "shipping_orders"
            )
            assert shipping["value"] == 30
            assert shipping["source_type"] == "system"
            completion = next(
                item
                for item in analytics["completion"]["items"]
                if item["code"] == "shipping_orders"
            )
            assert completion["state"] == "system"
            assert completion["label"] == "系统取数"
    finally:
        cleanup_test_month()


def test_shipping_system_sync_rejects_historical_month(monkeypatch) -> None:
    cleanup_test_month()
    monkeypatch.setattr(
        analytics_api,
        "current_business_month",
        lambda: date(2031, 4, 1),
    )

    def unexpected_source(_: date) -> dict[str, object]:
        raise AssertionError("历史月份不应再查询远程业务库")

    monkeypatch.setattr(analytics_api, "fetch_shipping_orders", unexpected_source)
    try:
        with TestClient(app) as client:
            login(client)
            preview = client.get(
                f"{settings.api_prefix}/analytics/details/shipping_orders/system-preview",
                params={"month": "2031-01"},
            )
            assert preview.status_code == 409
            assert "历史快照" in preview.json()["detail"]

            synced = client.post(
                f"{settings.api_prefix}/analytics/details/shipping_orders/system-sync",
                params={"month": "2031-01"},
            )
            assert synced.status_code == 409
            assert "历史快照" in synced.json()["detail"]

            details = client.get(
                f"{settings.api_prefix}/analytics/details/shipping_orders",
                params={"month": "2031-01"},
            )
            assert details.status_code == 200
            assert details.json()["snapshot"]["state"] == "historical"
            assert details.json()["snapshot"]["can_system_sync"] is False
    finally:
        cleanup_test_month()


def test_return_source_uses_monthly_eight_oclock_window(monkeypatch) -> None:
    captured: dict[str, object] = {}

    class FakeCursor:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return None

        def execute(self, query: str, params: tuple[object, ...]) -> None:
            captured["query"] = query
            captured["params"] = params

        def fetchall(self):
            return [
                {
                    "teamId": 7,
                    "teamName": "测试退货组",
                    "returnGoodsCount": 10,
                    "interceptChargeCount": 2,
                    "unusualChargeCount": 3,
                }
            ]

    class FakeConnection:
        def cursor(self):
            return FakeCursor()

        def close(self) -> None:
            captured["closed"] = True

    monkeypatch.setattr(
        analytics_sources.pymysql,
        "connect",
        lambda **_kwargs: FakeConnection(),
    )
    result = analytics_sources.fetch_return_items(date(2031, 2, 1))
    assert captured["params"] == (
        datetime(2031, 2, 1, 8),
        datetime(2031, 3, 1, 8),
        1,
    )
    assert "w.charge_num IS NULL AND w.total_num > 0" in str(captured["query"])
    assert "a.created_time >= p.start_time" in str(captured["query"])
    assert "a.created_time < p.end_time" in str(captured["query"])
    assert result["month_start"] == "2031-02-01 08:00:00"
    assert result["month_end"] == "2031-03-01 08:00:00"
    assert result["total"] == 15
    assert result["rows"][0]["退货件数合计"] == 15
    assert result["rows"][0]["数据退货占比"] == 100
    assert captured["closed"] is True


def test_return_system_preview_sync_and_monthly_total(monkeypatch) -> None:
    cleanup_return_test_month()
    monkeypatch.setattr(
        analytics_api,
        "current_business_month",
        lambda: date(2031, 4, 1),
    )

    def fake_return_source(selected: date) -> dict[str, object]:
        assert selected == date(2031, 2, 1)
        return {
            "source_name": "退货云端测试数据源",
            "month_start": "2031-02-01 08:00:00",
            "month_end": "2031-03-01 08:00:00",
            "columns": [
                "团队名称",
                "处理退货件数",
                "拦截件扣费件数",
                "异常件扣费件数",
                "退货件数合计",
                "数据退货占比",
            ],
            "rows": [
                {
                    "团队名称": "测试一组",
                    "处理退货件数": 10,
                    "拦截件扣费件数": 2,
                    "异常件扣费件数": 3,
                    "退货件数合计": 15,
                    "数据退货占比": 50,
                },
                {
                    "团队名称": "测试二组",
                    "处理退货件数": 5,
                    "拦截件扣费件数": 5,
                    "异常件扣费件数": 5,
                    "退货件数合计": 15,
                    "数据退货占比": 50,
                },
            ],
            "row_count": 2,
            "total": 30,
            "conditions": ["统计区间按月08:00左闭右开", "warehouse_id = 1"],
        }

    monkeypatch.setattr(analytics_api, "fetch_return_items", fake_return_source)
    try:
        with SessionLocal() as db:
            metric = db.scalar(
                select(MetricDefinition).where(MetricDefinition.code == "return_items")
            )
            assert metric is not None
            db.add(
                MonthlyMetric(
                    metric_id=metric.id,
                    month=date(2031, 1, 1),
                    value=20,
                    source_type="manual",
                )
            )
            db.commit()

        with TestClient(app) as client:
            login(client)
            preview = client.get(
                f"{settings.api_prefix}/analytics/details/return_items/system-preview",
                params={"month": "2031-02"},
            )
            assert preview.status_code == 200
            assert preview.json()["month_start"] == "2031-02-01 08:00:00"
            assert preview.json()["month_end"] == "2031-03-01 08:00:00"
            assert preview.json()["total"] == 30
            assert preview.json()["snapshot"]["state"] == "review"

            synced = client.post(
                f"{settings.api_prefix}/analytics/details/return_items/system-sync",
                params={"month": "2031-02"},
            )
            assert synced.status_code == 200
            assert synced.json()["total"] == 30

            details = client.get(
                f"{settings.api_prefix}/analytics/details/return_items",
                params={"month": "2031-02", "sort_order": "desc"},
            ).json()
            assert details["summary"]["return_items"] == 30
            assert details["snapshot"]["version_count"] == 1
            assert details["columns"] == [
                "团队名称",
                "处理退货件数",
                "拦截件扣费件数",
                "异常件扣费件数",
                "退货件数合计",
                "数据退货占比",
            ]
            assert details["rows"][0]["values"]["退货件数合计"] == 15

            searched = client.get(
                f"{settings.api_prefix}/analytics/details/return_items",
                params={"month": "2031-02", "search": "二"},
            ).json()
            assert searched["total"] == 1
            assert searched["rows"][0]["values"]["团队名称"] == "测试二组"

            exported = client.get(
                f"{settings.api_prefix}/analytics/details/return_items/export",
                params={"month": "2031-02", "search": "一"},
            )
            assert exported.status_code == 200
            export_book = load_workbook(BytesIO(exported.content), data_only=True)
            export_sheet = export_book["退货件数"]
            assert [cell.value for cell in export_sheet[1]] == details["columns"]
            assert export_sheet.max_row == 2
            assert export_sheet["A2"].value == "测试一组"
            assert export_sheet["E2"].value == 15

            analytics = client.get(
                f"{settings.api_prefix}/analytics", params={"month": "2031-02"}
            ).json()
            returns = next(
                item for item in analytics["metrics"] if item["code"] == "return_items"
            )
            assert returns["value"] == 30
            assert returns["previous_value"] == 20
            assert returns["change"] == 10
            assert returns["change_ratio"] == 50
            assert returns["source_type"] == "system"
    finally:
        cleanup_return_test_month()
