# ruff: noqa: E501
"""
客户对账账单拆分 V3.3
==================================================
【核心功能】
1. 按团队拆分账单，自动生成标准化Excel格式
2. 运费核算：全国均重/单票/新西1-3kg/新西1kg内加收费用计算
3. 专项统计：新西1kg订单数、1%占比判断、应结算单数核算
4. 文件命名：{合计金额}-{团队名}_{业务年月}_快递加收费.xlsx
5. 0费用过滤：普通团队合计金额为0不生成文件

【V2.4 架构整理】
  - 删除重复的 init_folder 函数，改从 utils 导入 ensure_folder
  - 样式常量改从 utils 导入，保持全项目Excel风格一致
  - 所有计费逻辑、千耀传媒专项、统计表完全保留V2.0框架

【可独立运行】
==================================================
"""

import math
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from app.domains.express import legacy_settings as settings
from app.domains.express.customer_special_rules import (
    DEFAULT_XIXI_1KG_UNIT_PRICE,
    ensure_special_rule_table,
)
from app.domains.express.utils import (
    BOLD_FONT,
    CENTER_ALIGN,
    FULL_BORDER,
    HEADER_FONT,
    LEFT_ALIGN,
    RIGHT_ALIGN,
    WRAP_CENTER_ALIGN,
    ensure_folder,
)

# ===================== 全局配置 =====================
SOURCE_FILE = os.path.join(settings.OUTPUT_FOLDER, settings.RESULT_FILE)
CONFIG_FILE = os.path.join(settings.CONFIG_FOLDER, "price_config.xlsx")
SAVE_DIR = os.path.join(settings.OUTPUT_FOLDER, "客户账单")

GROUP_COL = "所属团队"
FILTER_INVALID = True
TEST_TEAMS = []
SPECIAL_TEAM = "千耀传媒"

# 特殊客户：输出到 客户账单/特殊客户/ 子文件夹，0 元也强制生成
SPECIAL_FOLDER_TEAMS = ["汉尚华莲汉服", "MeftunStyle", "Vinnager&毅播"]
SPECIAL_FOLDER_DIR = os.path.join(SAVE_DIR, "特殊客户")

# ===================== 补充样式（split_bill专用）=====================
RED_BOLD_FONT = Font(bold=True, color="FF0000")

COL_WIDTH = {
    "A": 20,
    "B": 18,
    "C": 14,
    "D": 12,
    "E": 8,
    "F": 12,
    "G": 12,
    "H": 13,
    "I": 13,
    "L": 8,
    "M": 18,
    "N": 12,
    "O": 14,
    "P": 14,
    "Q": 12,
    "R": 15,
    "S": 15,
}
ROW_HEIGHT = 24
TEAM_CONFIG = {}


# ===================== 工具函数 =====================
def safe_file_name(raw_name: str) -> str:
    illegal_chars = ["/", "\\", ":", "*", "?", '"', "<", ">", "|"]
    name = str(raw_name)
    for char in illegal_chars:
        name = name.replace(char, "_")
    return name


def normalize_group_names(df: pd.DataFrame) -> pd.DataFrame:
    """清理团队名；缺失值保留为空，避免被转换成字符串 ``nan``。"""
    if GROUP_COL not in df.columns:
        raise KeyError(f"原始数据缺少列：{GROUP_COL}")
    df[GROUP_COL] = df[GROUP_COL].fillna("").astype(str).str.strip()
    return df


def load_team_config():
    global TEAM_CONFIG
    try:
        import pymysql

        conn = pymysql.connect(**settings.LOCAL_DB_CONFIG)
        with conn.cursor() as cur:
            ensure_special_rule_table(cur)
            cur.execute(
                "SELECT price.team_name, price.st_avg, price.zt_avg, "
                "price.st_extra, price.zt_extra, rule.xixi_1kg_unit_price "
                "FROM team_express_prices AS price "
                "LEFT JOIN team_special_rules AS rule "
                "ON rule.team_id = price.id "
                "ORDER BY price.seq"
            )
            rows = cur.fetchall()
        conn.close()
        TEAM_CONFIG = {
            r[0]: (
                float(r[1]),
                float(r[2]),
                float(r[3]),
                float(r[4]),
                float(r[5]) if r[5] is not None else DEFAULT_XIXI_1KG_UNIT_PRICE,
            )
            for r in rows
        }
        print(f"【配置】加载 {len(TEAM_CONFIG)} 个团队规则")
    except Exception as e:
        print(f"【错误】读取团队配置失败: {str(e)}")


def get_team_rule(team_name: str) -> tuple:
    return TEAM_CONFIG.get(str(team_name).strip(), (0.0, 0.0, 0.0, 0.0))[:4]


def get_xixi_1kg_unit_price(team_name: str) -> float:
    """未配置特殊价格时，保留原有的 10 元/单规则。"""
    rule = TEAM_CONFIG.get(str(team_name).strip(), ())
    if len(rule) > 4 and rule[4] is not None:
        return float(rule[4])
    return DEFAULT_XIXI_1KG_UNIT_PRICE


def set_worksheet_style(ws, max_row: int, max_col: int):
    for col, width in COL_WIDTH.items():
        ws.column_dimensions[col].width = width
    for row in range(1, max_row + 1):
        ws.row_dimensions[row].height = ROW_HEIGHT
    for row in range(1, max_row + 1):
        ws[f"B{row}"].number_format = "@"
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = FULL_BORDER
            cell.alignment = CENTER_ALIGN
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


# ===================== 核心汇总统计 =====================
def add_summary_area(ws, team_df: pd.DataFrame, team_name: str) -> float:
    st_avg_all, zt_avg_all, st_fee_all, zt_fee_all = get_team_rule(team_name)
    xixi_1kg_unit_price = get_xixi_1kg_unit_price(team_name)

    col_start = 12

    total_order = len(team_df)
    xixi_order = len(
        team_df[
            (team_df["结算重量"] < 1) & (team_df["目的省份"].str.contains("新疆|西藏", na=False))
        ]
    )
    threshold = total_order * 0.01
    is_over_flag = "是" if xixi_order >= threshold else "否"
    settle_order = math.ceil(xixi_order - threshold) if is_over_flag == "是" else 0
    r11_value = settle_order * xixi_1kg_unit_price

    # 全部客户统一使用按快递类型展示的汇总版式。
    # 仅全国均重按申通/中通拆分；其他项目保持原有“全部订单”取数口径。
    summary_headers = [
        "序号",
        "实际计算方式",
        "快递类型",
        "发货单量",
        "结算重量",
        "平均重量",
        "超出重量",
        "应付金额",
    ]
    calc_list = [
        ("全国均重", "申通"),
        ("全国均重", "中通"),
        ("单票", "全部"),
        ("新西1-3公斤", "全部"),
        ("新西1kg内（包1%）", "全部"),
        ("合计", ""),
    ]

    col_end = col_start + len(summary_headers) - 1

    # 汇总大标题
    ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
    title_cell = ws.cell(row=1, column=col_start)
    title_cell.value = "加收费汇总"
    title_cell.font = HEADER_FONT
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = CENTER_ALIGN
    title_cell.border = FULL_BORDER

    for idx, text in enumerate(summary_headers):
        cell = ws.cell(row=2, column=col_start + idx)
        cell.value = text
        cell.border = FULL_BORDER
        cell.alignment = CENTER_ALIGN
        cell.font = BOLD_FONT

    summary_rows = []
    total_fee_all = 0.0

    for idx, (calc_name, express_type) in enumerate(calc_list):
        seq = idx + 1
        order_cnt = total_weight = avg_weight = exceed_weight = row_fee = 0.0

        if calc_name != "合计":
            filter_df = team_df[team_df["实际计算方式"] == calc_name]
            if calc_name == "全国均重":
                filter_df = filter_df[filter_df["快递类型"] == express_type]
            order_cnt = len(filter_df)
            if order_cnt > 0:
                total_weight = round(filter_df["结算重量"].sum(), 2)
                avg_weight = round(total_weight / order_cnt, 2)

        if calc_name == "全国均重":
            use_avg = st_avg_all if express_type == "申通" else zt_avg_all
            use_fee = st_fee_all if express_type == "申通" else zt_fee_all
            exceed_weight = max(round(avg_weight - use_avg, 2), 0.0)
            single_fee = round((exceed_weight / 0.1) * use_fee, 2)
            row_fee = round(single_fee * order_cnt, 2)
            total_fee_all += row_fee
        elif calc_name in ("单票", "新西1-3公斤"):
            _fee = pd.to_numeric(filter_df["单票应付金额"], errors="coerce").fillna(0)
            row_fee = round(_fee.sum(), 2)
            total_fee_all += row_fee
        elif calc_name == "新西1kg内（包1%）":
            row_fee = r11_value
            total_fee_all += row_fee

        if calc_name == "新西1kg内（包1%）":
            row_data = [seq, calc_name, express_type, "", "", "", "", row_fee]
        elif calc_name == "合计":
            row_data = [seq, calc_name, "", total_order, "", "", "", "=SUM(S3:S7)"]
        else:
            row_data = [
                seq,
                calc_name,
                express_type,
                order_cnt,
                total_weight,
                avg_weight,
                exceed_weight,
                row_fee,
            ]

        if calc_name in ("单票", "新西1-3公斤", "新西1kg内（包1%）", "合计"):
            row_data[4] = row_data[5] = row_data[6] = ""

        summary_rows.append(row_data)

    total_fee_all = round(total_fee_all, 2)

    for row_idx, row_data in enumerate(summary_rows, start=3):
        for col_idx, val in enumerate(row_data):
            cell = ws.cell(row=row_idx, column=col_start + col_idx)
            cell.value = val
            cell.border = FULL_BORDER
            if col_idx == 0:
                cell.alignment = CENTER_ALIGN
            elif col_idx == 1:
                cell.alignment = LEFT_ALIGN
            else:
                cell.alignment = RIGHT_ALIGN
            if row_data[1] == "合计":
                if col_idx == 1:
                    cell.font = BOLD_FONT
                    cell.alignment = CENTER_ALIGN
                if col_idx == 7:
                    cell.font = RED_BOLD_FONT
        ws.row_dimensions[row_idx].height = ROW_HEIGHT

    # 专项统计表
    stat_row_t1, stat_row_t2, stat_row_data = 10, 11, 12
    stat_headers = [
        "序号",
        "实际计算方式",
        "快递类型",
        "总发货单量",
        "新西1kg内订单数",
        "是否超1%",
        "新西1kg内应结算单数",
        "新西1kg内应付金额",
    ]

    for idx in range(len(stat_headers)):
        col = col_start + idx
        ws.merge_cells(start_row=stat_row_t1, start_column=col, end_row=stat_row_t2, end_column=col)
        cell = ws.cell(row=stat_row_t1, column=col)
        cell.value = stat_headers[idx]
        cell.alignment = WRAP_CENTER_ALIGN
        cell.font = BOLD_FONT
        cell.border = FULL_BORDER

    stat_data = [
        "1",
        "新西1kg内（包1%）",
        "全部",
        total_order,
        xixi_order,
        is_over_flag,
        settle_order,
        r11_value,
    ]

    for idx, val in enumerate(stat_data):
        cell = ws.cell(row=stat_row_data, column=col_start + idx)
        cell.value = val
        cell.border = FULL_BORDER
        if idx == 0:
            cell.alignment = CENTER_ALIGN
        elif idx == 1:
            cell.alignment = LEFT_ALIGN
        else:
            cell.alignment = RIGHT_ALIGN

    for row in range(stat_row_t1, stat_row_data + 1):
        for col in range(col_start, col_start + len(stat_headers)):
            ws.cell(row=row, column=col).border = FULL_BORDER

    for r in [stat_row_t1, stat_row_t2, stat_row_data]:
        ws.row_dimensions[r].height = ROW_HEIGHT

    return total_fee_all


# ===================== 数据读取与文件拆分 =====================
def load_source_data() -> pd.DataFrame:
    if not os.path.exists(SOURCE_FILE):
        raise FileNotFoundError(f"原始数据文件不存在 {SOURCE_FILE}")
    df = pd.read_excel(SOURCE_FILE, engine="openpyxl")
    print(f"【数据】原始条数：{len(df)}")
    df = normalize_group_names(df)
    if FILTER_INVALID:
        df = df[(df[GROUP_COL] != "") & (df[GROUP_COL] != "未匹配")]
        print(f"【数据】有效条数：{len(df)}")
    df["运单号"] = df["运单号"].astype(str)
    return df


def split_by_group(df: pd.DataFrame):
    group_list = df.groupby(GROUP_COL)
    print(f"\n【分组】共识别 {len(group_list)} 个团队")
    print("===== 开始拆分账单 =====")

    for team_name, team_data in group_list:
        if TEST_TEAMS and team_name not in TEST_TEAMS:
            continue

        safe_team_name = safe_file_name(team_name)
        out_dir = SPECIAL_FOLDER_DIR if team_name in SPECIAL_FOLDER_TEAMS else SAVE_DIR
        ensure_folder(out_dir)
        temp_path = os.path.join(out_dir, f"temp_{safe_team_name}.xlsx")
        team_data.to_excel(temp_path, index=False, engine="openpyxl")

        wb = load_workbook(temp_path)
        ws = wb.active
        set_worksheet_style(ws, ws.max_row, ws.max_column)
        total_fee = add_summary_area(ws, team_data, team_name)
        wb.save(temp_path)
        wb.close()

        try:
            team_data["业务时间"] = pd.to_datetime(team_data["业务时间"])
            target_time = team_data.iloc[1 if len(team_data) >= 2 else 0]["业务时间"]
            year_month = target_time.strftime("%Y-%m")
        except Exception:
            year_month = "0000-00"

        if (
            abs(total_fee) < 1e-6
            and team_name != SPECIAL_TEAM
            and team_name not in SPECIAL_FOLDER_TEAMS
        ):
            if os.path.exists(temp_path):
                os.remove(temp_path)
            print(f"【跳过】{team_name}：合计应付为0")
            continue

        fee_str = f"{total_fee:.2f}"
        new_file_name = f"{fee_str}-{safe_team_name}_{year_month}_快递加收费.xlsx"
        new_path = os.path.join(out_dir, new_file_name)
        if os.path.exists(new_path):
            os.remove(new_path)
        os.rename(temp_path, new_path)
        print(f"✅ 生成完成：{new_file_name} | 单据数：{len(team_data)} | 合计：{total_fee}")


# ===================== 程序入口 =====================
def main():
    print("=" * 65)
    print("        客户对账账单拆分程序 V2.4 稳定版")
    print("=" * 65)
    try:
        ensure_folder(SAVE_DIR)
        ensure_folder(SPECIAL_FOLDER_DIR)
        load_team_config()
        source_df = load_source_data()
        split_by_group(source_df)
        print("\n" + "=" * 65)
        print("🎉 全部任务执行完毕")
        print("=" * 65)
    except Exception as e:
        print(f"\n❌ 运行异常：{str(e)}")
        raise


if __name__ == "__main__":
    main()
