"""FastAPI routes for the fully migrated express-reconciliation module."""

from __future__ import annotations

import hashlib
import json
import queue
import shutil
import zipfile
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

import pandas as pd
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.api.deps import require_permission
from app.core.config import settings
from app.db.session import get_db
from app.domains.express import legacy_settings
from app.models.operations import (
    ExpressCarrier,
    ExpressChargePrice,
    StoredFile,
    SystemSetting,
    TeamExpressPrice,
    TeamSpecialRule,
)
from app.models.user import User
from app.schemas.operations import (
    CarrierListInput,
    ExpressSettingsInput,
    TeamPriceListInput,
)
from app.services import express_jobs
from app.services.express_stats import (
    anomaly_frame,
    available_months,
    history_records,
    month_stats,
    output_folder,
    preview_rows,
    trend_data,
    unmatched_summary,
    validate_month,
)

router = APIRouter(prefix="/express", tags=["快递对账"])


def auth_user(user: User = Depends(require_permission("express.view"))) -> User:
    return user


@router.get("/overview")
def overview(_: User = Depends(auth_user)) -> dict[str, object]:
    months = available_months()
    selected = (
        legacy_settings.PROCESS_MONTH
        if legacy_settings.PROCESS_MONTH in months
        else (months[0] if months else legacy_settings.PROCESS_MONTH)
    )
    return {
        "process_month": legacy_settings.PROCESS_MONTH,
        "selected_month": selected,
        "months": months,
        "stats": month_stats(selected),
        "trend": trend_data(),
        "recent_runs": history_records()[:6],
        "job": dict(express_jobs.state),
    }


@router.post("/upload")
def upload_bill(
    month: str = Form(default=legacy_settings.PROCESS_MONTH),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("express.run")),
) -> dict[str, object]:
    validate_month(month)
    filename = Path(file.filename or "").name
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 账单文件")
    destination = settings.storage_path / "data" / month / filename
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("wb") as output:
        shutil.copyfileobj(file.file, output)
    if destination.stat().st_size > 100 * 1024 * 1024:
        destination.unlink(missing_ok=True)
        raise HTTPException(status_code=413, detail="单个文件不能超过 100 MB")

    relative = destination.relative_to(settings.storage_path).as_posix()
    target = db.scalar(select(StoredFile).where(StoredFile.relative_path == relative))
    digest = hashlib.sha256(destination.read_bytes()).hexdigest()
    if target is None:
        target = StoredFile(
            category="express_input",
            period=month,
            original_name=filename,
            relative_path=relative,
            size_bytes=destination.stat().st_size,
            sha256=digest,
            source="new_system",
        )
        db.add(target)
    else:
        target.size_bytes = destination.stat().st_size
        target.sha256 = digest
    db.commit()
    return {"ok": True, "filename": filename, "size_bytes": destination.stat().st_size}


@router.post("/run")
def run_reconciliation(
    user: User = Depends(require_permission("express.run")),
) -> dict[str, object]:
    ok, message = express_jobs.start_reconciliation(user.id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=message)
    return {"ok": True, "message": message, "period": legacy_settings.PROCESS_MONTH}


@router.get("/status")
def job_status(_: User = Depends(require_permission("express.run"))) -> dict[str, object]:
    return dict(express_jobs.state)


@router.get("/logs")
def stream_logs(_: User = Depends(require_permission("express.run"))) -> StreamingResponse:
    def generate():
        while True:
            try:
                message = express_jobs.log_queue.get(timeout=20)
            except queue.Empty:
                message = "__PING__"
            yield f"data: {json.dumps(message, ensure_ascii=False)}\n\n"
            if message == "__DONE__":
                break

    return StreamingResponse(generate(), media_type="text/event-stream")


@router.get("/history")
def history(_: User = Depends(auth_user)) -> list[dict[str, object]]:
    return history_records()


@router.get("/stats/{month}")
def stats(month: str, _: User = Depends(auth_user)) -> dict[str, object]:
    try:
        return month_stats(month)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.get("/trend")
def trend(_: User = Depends(auth_user)) -> list[dict[str, object]]:
    return trend_data()


@router.get("/unmatched/{month}")
def unmatched(month: str, _: User = Depends(auth_user)) -> dict[str, object]:
    result = unmatched_summary(month)
    if not result["ok"]:
        raise HTTPException(status_code=404, detail=result["msg"])
    return result


@router.get("/preview/{month}")
def preview(
    month: str,
    page: int = Query(default=1, ge=1),
    size: int = Query(default=50, ge=10, le=200),
    filter_: str = Query(default="all", alias="filter"),
    keyword: str = "",
    _: User = Depends(auth_user),
) -> dict[str, object]:
    if filter_ not in {"all", "matched", "unmatched", "single", "average"}:
        raise HTTPException(status_code=400, detail="不支持的筛选条件")
    result = preview_rows(month, page, size, filter_, keyword.strip())
    if not result["ok"]:
        raise HTTPException(status_code=404, detail=f"{month} 对账结果文件不存在")
    return result


@router.get("/download/{month}")
def download_month(
    month: str, _: User = Depends(require_permission("express.download"))
) -> FileResponse:
    folder = output_folder(month)
    if not folder.exists():
        raise HTTPException(status_code=404, detail="该月份没有结果文件")
    zip_path = settings.storage_path / "output" / f"{month}_结果.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in folder.rglob("*"):
            if path.is_file():
                archive.write(path, path.relative_to(folder.parent))
    return FileResponse(zip_path, filename=f"{month}_对账结果.zip")


@router.get("/anomalies/{month}/download")
def download_anomalies(
    month: str, _: User = Depends(require_permission("express.download"))
) -> StreamingResponse:
    frame = anomaly_frame(month)
    if frame.empty:
        raise HTTPException(status_code=404, detail=f"{month} 未发现任何异常运单")
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False)
        sheet = writer.book.active
        headers = [cell.value for cell in sheet[1]]
        type_column = headers.index("异常类型") + 1
        reason_column = headers.index("异常原因说明") + 1
        fill_high = PatternFill("solid", fgColor="4D1010")
        fill_mid = PatternFill("solid", fgColor="4D3010")
        fill_head = PatternFill("solid", fgColor="1E2235")
        for cell in sheet[1]:
            cell.fill = fill_head
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows(min_row=2):
            type_cell = row[type_column - 1]
            types = str(type_cell.value or "").split("/")
            type_cell.fill = fill_high if "重量异常" in types else fill_mid
            type_cell.font = Font(bold=True, color="FFFFFF", size=11)
            type_cell.alignment = Alignment(horizontal="center", vertical="center")
            row[reason_column - 1].alignment = Alignment(wrap_text=True, vertical="top")
        for index, header in enumerate(headers, start=1):
            if header == "运单号":
                width = 22
            elif header == "异常类型":
                width = 18
            elif header == "异常原因说明":
                width = 55
            elif header in ("目的省份", "目的城市", "快递类型", "所属团队"):
                width = 14
            else:
                width = 12
            sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
        sheet.freeze_panes = "A2"
    buffer.seek(0)
    filename = quote(f"{month}_异常运单_{len(frame)}条.xlsx")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.get("/config/prices")
def get_prices(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("express.configure")),
) -> dict[str, object]:
    price_file = settings.storage_path / "config" / "price_config.xlsx"
    if not price_file.exists():
        raise HTTPException(status_code=404, detail="报价表不存在")
    result: dict[str, object] = {"shentong": [], "zhongtong": [], "charge": []}
    for sheet, key in (("申通报价", "shentong"), ("中通报价", "zhongtong")):
        frame = pd.read_excel(price_file, sheet_name=sheet)
        rows = []
        for _, row in frame.iterrows():
            if pd.isna(row.iloc[0]) or not str(row.iloc[0]).strip():
                continue
            rows.append(
                {
                    "province": str(row.iloc[0]).strip(),
                    "fee_3kg": float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0,
                    "fee_over3kg": float(row.iloc[3]) if pd.notna(row.iloc[3]) else 0,
                    "unit_price": float(row.iloc[4]) if pd.notna(row.iloc[4]) else 0,
                }
            )
        result[key] = rows
    result["charge"] = [
        {"type": row.express_type, "price": float(row.charge_price)}
        for row in db.scalars(select(ExpressChargePrice).order_by(ExpressChargePrice.id))
    ]
    return result


@router.put("/config/prices")
def save_prices(
    payload: dict[str, list[dict[str, object]]],
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("express.configure")),
) -> dict[str, object]:
    price_file = settings.storage_path / "config" / "price_config.xlsx"
    if not price_file.exists():
        raise HTTPException(status_code=404, detail="报价表不存在")
    workbook = load_workbook(price_file)
    for key, sheet_name in (("shentong", "申通报价"), ("zhongtong", "中通报价")):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        for index, item in enumerate(payload.get(key, []), start=2):
            sheet.cell(index, 1).value = item.get("province")
            sheet.cell(index, 2).value = item.get("fee_3kg", 0)
            sheet.cell(index, 4).value = item.get("fee_over3kg", 0)
            sheet.cell(index, 5).value = item.get("unit_price", 0)
    for item in payload.get("charge", []):
        express_type = str(item.get("type", ""))
        target = db.scalar(
            select(ExpressChargePrice).where(ExpressChargePrice.express_type == express_type)
        )
        if target:
            target.charge_price = item.get("price", 0)
    workbook.save(price_file)
    db.commit()
    return {"ok": True, "message": "报价配置已保存"}


@router.get("/config/carriers")
def get_carriers(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("express.configure")),
) -> list[dict[str, object]]:
    rows = db.scalars(select(ExpressCarrier).order_by(ExpressCarrier.sort_order)).all()
    return [
        {
            "id": row.id,
            "name": row.name,
            "identify_column": row.identify_column,
            "enabled": row.enabled,
        }
        for row in rows
    ]


@router.put("/config/carriers")
def save_carriers(
    payload: CarrierListInput,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("express.configure")),
) -> dict[str, object]:
    db.execute(delete(ExpressCarrier))
    rows = []
    for index, item in enumerate(payload.carriers, start=1):
        db.add(
            ExpressCarrier(
                name=item.name.strip(),
                identify_column=item.identify_column.strip(),
                enabled=item.enabled,
                sort_order=index,
            )
        )
        rows.append(item.model_dump())
    config_path = settings.storage_path / "config" / "express_config.json"
    config_path.write_text(
        json.dumps({"express_list": rows}, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    db.commit()
    return {"ok": True, "message": f"已保存 {len(rows)} 家快递配置"}


@router.get("/config/settings")
def get_settings(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("express.configure")),
) -> dict[str, object]:
    before = int(db.get(SystemSetting, "express.extend_days_before").value)
    after = int(db.get(SystemSetting, "express.extend_days_after").value)
    first = datetime.now().replace(day=1)
    last = first - timedelta(days=1)
    start = last.replace(day=1) - timedelta(days=before)
    end = last + timedelta(days=after)
    return {
        "extend_days_before": before,
        "extend_days_after": after,
        "process_month": legacy_settings.PROCESS_MONTH,
        "sql_start_date": start.strftime("%Y-%m-%d 00:00:00"),
        "sql_end_date": end.strftime("%Y-%m-%d 23:59:59"),
    }


@router.put("/config/settings")
def save_settings(
    payload: ExpressSettingsInput,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("express.configure")),
) -> dict[str, object]:
    values = {
        "express.extend_days_before": payload.extend_days_before,
        "express.extend_days_after": payload.extend_days_after,
    }
    for key, value in values.items():
        setting = db.get(SystemSetting, key)
        if setting is None:
            setting = SystemSetting(key=key, value=str(value))
            db.add(setting)
        else:
            setting.value = str(value)
    override = {
        "SQL_EXTEND_DAYS_BEFORE": payload.extend_days_before,
        "SQL_EXTEND_DAYS_AFTER": payload.extend_days_after,
    }
    (settings.storage_path / "config" / "settings_override.json").write_text(
        json.dumps(override, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    db.commit()
    return {"ok": True, "message": "运行参数已保存"}


@router.get("/config/customers")
def get_customers(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("express.configure")),
) -> list[dict[str, object]]:
    rows = db.scalars(select(TeamExpressPrice).order_by(TeamExpressPrice.seq)).all()
    return [
        {
            "team": row.team_name,
            "st_fee": float(row.st_fee),
            "st3": float(row.st_avg),
            "st01": float(row.st_extra),
            "zt_fee": float(row.zt_fee),
            "zt3": float(row.zt_avg),
            "zt01": float(row.zt_extra),
            "xixi_1kg_unit_price": (
                float(row.special_rule.xixi_1kg_unit_price) if row.special_rule else None
            ),
            "special_note": row.special_rule.special_note if row.special_rule else "",
        }
        for row in rows
    ]


@router.put("/config/customers")
def save_customers(
    payload: TeamPriceListInput,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("express.configure")),
) -> dict[str, object]:
    db.execute(delete(TeamSpecialRule))
    db.execute(delete(TeamExpressPrice))
    for index, item in enumerate(payload.rows, start=1):
        team = TeamExpressPrice(
            seq=index,
            team_name=item.team.strip(),
            st_fee=item.st_fee,
            st_avg=item.st3,
            st_extra=item.st01,
            zt_fee=item.zt_fee,
            zt_avg=item.zt3,
            zt_extra=item.zt01,
        )
        db.add(team)
        db.flush()
        if item.xixi_1kg_unit_price is not None or item.special_note.strip():
            db.add(
                TeamSpecialRule(
                    team_id=team.id,
                    xixi_1kg_unit_price=item.xixi_1kg_unit_price or 10,
                    special_note=item.special_note.strip() or None,
                )
            )
    db.commit()
    return {"ok": True, "message": f"已保存 {len(payload.rows)} 条客户配置"}
