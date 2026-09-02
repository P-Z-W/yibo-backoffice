"""Fast, two-level warehouse reimbursement workflow APIs."""

from __future__ import annotations

from datetime import UTC, date, datetime
from decimal import Decimal
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from urllib.parse import quote
from uuid import uuid4

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import delete, func, or_, select
from sqlalchemy.orm import Session, selectinload

from app.api.deps import (
    get_permission_scope,
    require_any_permission,
    require_permission,
    user_has_role,
)
from app.core.config import settings
from app.db.session import get_db
from app.models.operations import SystemSetting
from app.models.reimbursement import (
    Reimbursement,
    ReimbursementApproval,
    ReimbursementAttachment,
    ReimbursementEntity,
    ReimbursementInvoice,
    ReimbursementItem,
)
from app.models.user import User
from app.schemas.reimbursement import (
    ReimbursementActionInput,
    ReimbursementConfigInput,
    ReimbursementEntityInput,
    ReimbursementExportInput,
    ReimbursementInput,
    ReimbursementInvoiceInput,
)
from app.services.invoice_ocr import (
    InvoiceOcrError,
    InvoiceOcrNotConfiguredError,
    recognize_invoice,
)

router = APIRouter(prefix="/reimbursements", tags=["报销"])

FINANCE_SETTING_KEY = "reimbursement_finance_approval"
TEAMS = ["发货组", "退货组"]
EXPENSE_CATEGORIES = [
    "临时运费",
    "快递垫付款",
    "退件运费",
    "退货取件垫付款",
    "装卸搬运费",
    "包材临时采购",
    "补发换货垫付款",
    "检测维修费",
    "交通费",
    "加班餐费",
    "办公用品",
    "其他",
]
EDITABLE_STATUSES = {"draft", "returned"}
SUBMITTED_STATUSES = {"pending_supervisor", "pending_finance"}
ALLOWED_UPLOAD_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp", ".pdf", ".xlsx"}
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_BATCH_ROWS = 500
MAX_BATCH_CLAIMS = 100
BATCH_TEMPLATE_PATH = (
    Path(__file__).resolve().parents[2] / "assets" / "reimbursement_batch_import_template.xlsx"
)


def now() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def finance_approval_enabled(db: Session) -> bool:
    setting = db.get(SystemSetting, FINANCE_SETTING_KEY)
    return bool(setting and setting.value.strip().lower() in {"1", "true", "yes"})


def serialize_entity(entity: ReimbursementEntity) -> dict[str, object]:
    return {
        "id": entity.id,
        "name": entity.name,
        "tax_number": entity.tax_number,
        "is_default": entity.is_default,
        "is_active": entity.is_active,
    }


def active_entities(db: Session) -> list[dict[str, object]]:
    entities = db.scalars(
        select(ReimbursementEntity)
        .where(ReimbursementEntity.is_active.is_(True))
        .order_by(ReimbursementEntity.is_default.desc(), ReimbursementEntity.name)
    ).all()
    return [serialize_entity(entity) for entity in entities]


def reimbursement_config_payload(db: Session) -> dict[str, object]:
    return {
        "finance_approval_enabled": finance_approval_enabled(db),
        "teams": TEAMS,
        "expense_categories": EXPENSE_CATEGORIES,
        "entities": active_entities(db),
        "invoice_ocr_available": settings.invoice_ocr_available,
        "invoice_ocr_provider": settings.invoice_ocr_provider,
    }


def visible_statement(db: Session, user: User):
    statement = select(Reimbursement).options(
        selectinload(Reimbursement.items),
        selectinload(Reimbursement.attachments).selectinload(ReimbursementAttachment.invoice),
        selectinload(Reimbursement.approval_records),
    )
    scope = get_permission_scope(db, user, "reimbursement.view")
    if scope == "all":
        return statement
    if scope == "team":
        return statement.where(
            or_(Reimbursement.team == user.team, Reimbursement.applicant_id == user.id)
        )
    return statement.where(Reimbursement.applicant_id == user.id)


def get_visible_claim(db: Session, claim_id: int, user: User) -> Reimbursement:
    claim = db.scalar(visible_statement(db, user).where(Reimbursement.id == claim_id))
    if claim is None:
        raise HTTPException(status_code=404, detail="报销单不存在或无权查看")
    return claim


def can_edit(db: Session, claim: Reimbursement, user: User) -> bool:
    scope = get_permission_scope(db, user, "reimbursement.create")
    if claim.status not in EDITABLE_STATUSES or scope is None:
        return False
    if scope == "all":
        return True
    if scope == "team":
        return claim.team == user.team or claim.applicant_id == user.id
    return claim.applicant_id == user.id


def can_approve(db: Session, claim: Reimbursement, user: User) -> bool:
    if claim.status == "pending_supervisor":
        scope = get_permission_scope(db, user, "reimbursement.approve_supervisor")
        return bool(
            scope
            and (user_has_role(db, user, "admin") or user.id != claim.applicant_id)
            and (scope == "all" or (scope == "team" and user.team == claim.team))
        )
    if claim.status == "pending_finance":
        scope = get_permission_scope(db, user, "reimbursement.approve_finance")
        return bool(scope and (scope == "all" or (scope == "team" and user.team == claim.team)))
    return False


def display_status(claim: Reimbursement) -> str:
    labels = {
        "draft": "草稿",
        "pending_supervisor": "待主管审批",
        "pending_finance": "待财务审批",
        "returned": "已退回",
        "approved": "已审批",
        "withdrawn": "已撤回",
    }
    return labels.get(claim.status, claim.status)


def invalidate_export(claim: Reimbursement) -> None:
    """Mark a previously exported claim as changed without affecting its workflow."""
    claim.exported = False
    claim.exported_at = None
    claim.export_batch = None


def serialize_invoice(invoice: ReimbursementInvoice) -> dict[str, object]:
    return {
        "id": invoice.id,
        "recognition_status": invoice.recognition_status,
        "recognition_provider": invoice.recognition_provider,
        "recognition_message": invoice.recognition_message or "",
        "recognized_entity_name": invoice.recognized_entity_name or "",
        "recognized_tax_number": invoice.recognized_tax_number or "",
        "recognized_amount": (
            float(invoice.recognized_amount) if invoice.recognized_amount is not None else None
        ),
        "entity_name": invoice.final_entity_name or "",
        "tax_number": invoice.final_tax_number or "",
        "amount": float(invoice.final_amount) if invoice.final_amount is not None else None,
        "invoice_code": invoice.invoice_code or "",
        "invoice_number": invoice.invoice_number or "",
        "invoice_date": invoice.invoice_date.isoformat() if invoice.invoice_date else "",
        "manually_edited": invoice.manually_edited,
        "recognized_at": invoice.recognized_at.isoformat() if invoice.recognized_at else "",
    }


def serialize_attachment(
    attachment: ReimbursementAttachment, *, duplicate: bool = False
) -> dict[str, object]:
    return {
        "id": attachment.id,
        "original_name": attachment.original_name,
        "content_type": attachment.content_type,
        "size_bytes": attachment.size_bytes,
        "document_type": attachment.document_type,
        "url": f"{settings.api_prefix}/reimbursements/attachments/{attachment.id}",
        "created_at": attachment.created_at.isoformat() if attachment.created_at else "",
        "duplicate": duplicate,
        "invoice": serialize_invoice(attachment.invoice) if attachment.invoice else None,
    }


def serialize_claim(
    db: Session, claim: Reimbursement, user: User, *, detail: bool = False
) -> dict[str, object]:
    items = [
        {
            "id": item.id,
            "expense_date": item.expense_date.isoformat(),
            "category": item.category,
            "amount": float(item.amount),
            "related_number": item.related_number or "",
            "description": item.description or "",
        }
        for item in claim.items
    ]
    categories = []
    for item in claim.items:
        if item.category not in categories:
            categories.append(item.category)
    summary = "、".join(categories[:2]) or "未填写费用"
    if len(categories) > 2:
        summary += "等"
    summary += f" {len(items)} 项"
    invoices = [item.invoice for item in claim.attachments if item.invoice is not None]
    invoice_amount = sum(
        (invoice.final_amount for invoice in invoices if invoice.final_amount is not None),
        start=Decimal("0"),
    )
    invoice_issues = sum(
        invoice.recognition_status in {"failed", "unconfigured", "needs_review"}
        or bool(
            invoice.final_entity_name
            and claim.entity_name
            and invoice.final_entity_name != claim.entity_name
        )
        or bool(
            invoice.final_tax_number
            and claim.tax_number
            and invoice.final_tax_number != claim.tax_number
        )
        for invoice in invoices
    )
    payload: dict[str, object] = {
        "id": claim.id,
        "number": claim.number,
        "applicant_id": claim.applicant_id,
        "applicant_name": claim.applicant_name,
        "team": claim.team,
        "entity_name": claim.entity_name,
        "tax_number": claim.tax_number,
        "status": claim.status,
        "status_label": display_status(claim),
        "total_amount": float(claim.total_amount),
        "item_count": len(items),
        "item_summary": summary,
        "attachment_count": len(claim.attachments),
        "invoice_count": len(invoices),
        "invoice_amount": float(invoice_amount),
        "invoice_amount_difference": float(claim.total_amount - invoice_amount),
        "invoice_issue_count": invoice_issues,
        "note": claim.note or "",
        "finance_approval_required": claim.finance_approval_required,
        "exported": claim.exported,
        "exported_at": claim.exported_at.isoformat() if claim.exported_at else "",
        "export_batch": claim.export_batch or "",
        "submitted_at": claim.submitted_at.isoformat() if claim.submitted_at else "",
        "created_at": claim.created_at.isoformat() if claim.created_at else "",
        "updated_at": claim.updated_at.isoformat() if claim.updated_at else "",
        "can_edit": can_edit(db, claim, user),
        "can_approve": can_approve(db, claim, user),
    }
    if detail:
        payload["items"] = items
        payload["attachments"] = [serialize_attachment(item) for item in claim.attachments]
        payload["approval_records"] = [
            {
                "id": record.id,
                "actor_name": record.actor_name,
                "actor_role": record.actor_role,
                "action": record.action,
                "from_status": record.from_status,
                "to_status": record.to_status,
                "comment": record.comment or "",
                "created_at": record.created_at.isoformat() if record.created_at else "",
            }
            for record in claim.approval_records
        ]
    return payload


def replace_items(claim: Reimbursement, payload: ReimbursementInput) -> None:
    claim.applicant_name = payload.applicant_name
    claim.team = payload.team
    claim.entity_name = payload.entity_name
    claim.tax_number = payload.tax_number
    claim.note = payload.note or None
    claim.items.clear()
    total = Decimal("0")
    for index, row in enumerate(payload.items):
        total += row.amount
        claim.items.append(
            ReimbursementItem(
                expense_date=row.expense_date,
                category=row.category,
                amount=row.amount,
                related_number=row.related_number or None,
                description=row.description or None,
                sort_order=index,
            )
        )
    claim.total_amount = total.quantize(Decimal("0.01"))


def validate_for_submit(claim: Reimbursement) -> None:
    if not claim.entity_name:
        raise HTTPException(status_code=422, detail="请选择或填写报销主体")
    if not claim.tax_number:
        raise HTTPException(status_code=422, detail="请填写报销主体税号")
    if not claim.items:
        raise HTTPException(status_code=422, detail="至少填写一条费用明细")
    invalid = [index + 1 for index, item in enumerate(claim.items) if item.amount <= 0]
    if invalid:
        raise HTTPException(status_code=422, detail=f"第 {invalid[0]} 条明细的金额必须大于 0")


def add_action(
    claim: Reimbursement,
    user: User,
    action: str,
    from_status: str,
    to_status: str,
    comment: str = "",
) -> None:
    claim.approval_records.append(
        ReimbursementApproval(
            actor_id=user.id,
            actor_name=user.display_name,
            actor_role=user.role,
            action=action,
            from_status=from_status,
            to_status=to_status,
            comment=comment or None,
        )
    )


def generate_number(db: Session) -> str:
    prefix = f"BX{date.today():%Y%m%d}"
    last = db.scalar(
        select(Reimbursement.number)
        .where(Reimbursement.number.like(f"{prefix}%"))
        .order_by(Reimbursement.number.desc())
        .limit(1)
    )
    sequence = int(last[-3:]) + 1 if last else 1
    return f"{prefix}{sequence:03d}"


def parse_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip().replace("/", "-")
    if not text:
        return date.today()
    for pattern in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y%m%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError("日期格式应为 YYYY-MM-DD")


def parse_batch_workbook(
    contents: bytes,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    """Parse grouped reimbursement rows from the standard batch workbook."""
    try:
        workbook = load_workbook(BytesIO(contents), data_only=True, read_only=True)
        sheet = workbook["批量导入"] if "批量导入" in workbook.sheetnames else workbook.active
        rows = sheet.iter_rows(values_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="无法读取该 Excel 文件") from exc

    aliases = {
        "报销分组": "group_key",
        "报销批次": "group_key",
        "单据分组": "group_key",
        "分组": "group_key",
        "报销人": "applicant_name",
        "申请人": "applicant_name",
        "所属组": "team",
        "组别": "team",
        "主体": "entity_name",
        "报销主体": "entity_name",
        "购买方名称": "entity_name",
        "税号": "tax_number",
        "纳税人识别号": "tax_number",
        "购买方税号": "tax_number",
        "费用日期": "expense_date",
        "日期": "expense_date",
        "费用类别": "category",
        "类别": "category",
        "金额": "amount",
        "关联单号": "related_number",
        "单号": "related_number",
        "费用说明": "description",
        "说明": "description",
        "整单备注": "note",
        "报销备注": "note",
    }
    required = {
        "group_key",
        "applicant_name",
        "team",
        "entity_name",
        "tax_number",
        "expense_date",
        "category",
        "amount",
    }
    mapping: dict[int, str] = {}
    header_row = 0
    for row_number, values in enumerate(rows, start=1):
        headers = [
            str(value or "").replace("*", "").replace("（必填）", "").strip() for value in values
        ]
        candidate = {index: aliases[name] for index, name in enumerate(headers) if name in aliases}
        if required.issubset(candidate.values()):
            mapping = candidate
            header_row = row_number
            break
        if row_number >= 10:
            break
    if not mapping:
        raise HTTPException(
            status_code=422,
            detail="未找到标准表头，请使用系统提供的批量导入模板",
        )

    grouped: dict[str, dict[str, object]] = {}
    errors: list[dict[str, object]] = []
    data_rows = 0
    for excel_row, values in enumerate(rows, start=header_row + 1):
        raw = {
            field: values[index] if index < len(values) else None
            for index, field in mapping.items()
        }
        if not any(value not in {None, ""} for value in raw.values()):
            continue
        data_rows += 1
        if data_rows > MAX_BATCH_ROWS:
            errors.append({"row": excel_row, "message": f"单次最多导入 {MAX_BATCH_ROWS} 行"})
            break
        try:
            group_key = str(raw.get("group_key") or "").strip()
            applicant_name = str(raw.get("applicant_name") or "").strip()
            team = str(raw.get("team") or "").strip()
            entity_name = str(raw.get("entity_name") or "").strip()
            tax_number = "".join(str(raw.get("tax_number") or "").upper().split())
            category = str(raw.get("category") or "").strip()
            note = str(raw.get("note") or "").strip()
            if not group_key:
                raise ValueError("报销分组不能为空")
            if len(group_key) > 50:
                raise ValueError("报销分组不能超过 50 个字")
            if not applicant_name:
                raise ValueError("报销人不能为空")
            if len(applicant_name) > 80:
                raise ValueError("报销人不能超过 80 个字")
            if team not in TEAMS:
                raise ValueError("所属组只能填写发货组或退货组")
            if not entity_name:
                raise ValueError("主体不能为空")
            if not tax_number:
                raise ValueError("税号不能为空")
            if not category:
                raise ValueError("费用类别不能为空")
            if len(category) > 60:
                raise ValueError("费用类别不能超过 60 个字")
            if len(note) > 1000:
                raise ValueError("整单备注不能超过 1000 个字")
            amount = Decimal(str(raw.get("amount") or "0").replace(",", "")).quantize(
                Decimal("0.01")
            )
            if amount <= 0:
                raise ValueError("金额必须大于 0")
            related_number = str(raw.get("related_number") or "").strip()
            description = str(raw.get("description") or "").strip()
            if len(related_number) > 100:
                raise ValueError("关联单号不能超过 100 个字")
            if len(description) > 255:
                raise ValueError("费用说明不能超过 255 个字")
            item = {
                "expense_date": parse_date(raw.get("expense_date")),
                "category": category,
                "amount": amount,
                "related_number": related_number,
                "description": description,
            }
        except (ValueError, TypeError, ArithmeticError) as exc:
            errors.append({"row": excel_row, "message": str(exc)})
            continue

        group = grouped.get(group_key)
        if group is None:
            if len(grouped) >= MAX_BATCH_CLAIMS:
                errors.append(
                    {"row": excel_row, "message": f"单次最多生成 {MAX_BATCH_CLAIMS} 张报销单"}
                )
                continue
            group = {
                "group_key": group_key,
                "applicant_name": applicant_name,
                "team": team,
                "entity_name": entity_name,
                "tax_number": tax_number,
                "note": note,
                "items": [],
                "source_rows": [],
            }
            grouped[group_key] = group
        elif (
            group["applicant_name"] != applicant_name
            or group["team"] != team
            or group["entity_name"] != entity_name
            or group["tax_number"] != tax_number
        ):
            errors.append(
                {
                    "row": excel_row,
                    "message": f"分组 {group_key} 的报销人、所属组、主体和税号必须保持一致",
                }
            )
            continue
        elif note and group["note"] and group["note"] != note:
            errors.append({"row": excel_row, "message": f"分组 {group_key} 的整单备注必须保持一致"})
            continue
        elif note and not group["note"]:
            group["note"] = note
        group_items = group["items"]
        group_rows = group["source_rows"]
        assert isinstance(group_items, list)
        assert isinstance(group_rows, list)
        group_items.append(item)
        group_rows.append(excel_row)

    return list(grouped.values()), errors


def validate_batch_groups(
    db: Session, groups: list[dict[str, object]], user: User
) -> list[dict[str, object]]:
    """Attach applicant ids and return permission/account validation errors."""
    names = {str(group["applicant_name"]) for group in groups}
    users = db.scalars(
        select(User).where(User.display_name.in_(names), User.is_active.is_(True))
    ).all()
    user_map = {item.display_name: item for item in users}
    scope = get_permission_scope(db, user, "reimbursement.create")
    errors: list[dict[str, object]] = []
    for group in groups:
        group_key = str(group["group_key"])
        applicant = user_map.get(str(group["applicant_name"]))
        issues: list[str] = []
        if applicant is None:
            issues.append("报销人账号不存在或已停用")
        else:
            group["applicant_id"] = applicant.id
            if applicant.team and applicant.team != group["team"]:
                issues.append(f"账号所属组为{applicant.team}，与模板不一致")
            if scope == "self" and applicant.id != user.id:
                issues.append("当前账号只能导入自己的报销")
            elif scope == "team" and applicant.id != user.id and group["team"] != user.team:
                issues.append("当前账号只能导入本组报销")
            elif scope is None:
                issues.append("当前账号无填报权限")
        group["issues"] = issues
        group["valid"] = not issues
        for issue in issues:
            errors.append({"group": group_key, "message": issue})
    return errors


def batch_preview_payload(groups: list[dict[str, object]]) -> list[dict[str, object]]:
    preview: list[dict[str, object]] = []
    for group in groups:
        items = group["items"]
        assert isinstance(items, list)
        preview.append(
            {
                "group_key": group["group_key"],
                "applicant_name": group["applicant_name"],
                "team": group["team"],
                "entity_name": group["entity_name"],
                "tax_number": group["tax_number"],
                "note": group["note"],
                "item_count": len(items),
                "total_amount": float(sum((item["amount"] for item in items), start=Decimal("0"))),
                "valid": group.get("valid", True),
                "issues": group.get("issues", []),
            }
        )
    return preview


@router.get("")
def list_reimbursements(
    view: str = "all",
    team: str = "",
    keyword: str = "",
    start_date: date | None = None,
    end_date: date | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("reimbursement.view")),
) -> dict[str, object]:
    base = visible_statement(db, user)
    all_rows = db.scalars(base.order_by(Reimbursement.created_at.desc())).unique().all()

    statement = visible_statement(db, user)
    if view == "mine":
        approval_filters = []
        supervisor_scope = get_permission_scope(db, user, "reimbursement.approve_supervisor")
        if supervisor_scope:
            supervisor_filter = Reimbursement.status == "pending_supervisor"
            if supervisor_scope == "team":
                supervisor_filter = supervisor_filter & (Reimbursement.team == user.team)
            approval_filters.append(supervisor_filter)
        finance_scope = get_permission_scope(db, user, "reimbursement.approve_finance")
        if finance_scope:
            finance_filter = Reimbursement.status == "pending_finance"
            if finance_scope == "team":
                finance_filter = finance_filter & (Reimbursement.team == user.team)
            approval_filters.append(finance_filter)
        if approval_filters:
            statement = statement.where(or_(*approval_filters))
        else:
            statement = statement.where(Reimbursement.applicant_id == user.id)
    elif view == "pending_export":
        statement = statement.where(Reimbursement.exported.is_(False))
    elif view == "exported":
        statement = statement.where(Reimbursement.exported.is_(True))
    elif view and view not in {"all", "mine"}:
        statement = statement.where(Reimbursement.status == view)
    if team:
        statement = statement.where(Reimbursement.team == team)
    if keyword.strip():
        like = f"%{keyword.strip()}%"
        statement = statement.where(
            or_(
                Reimbursement.number.like(like),
                Reimbursement.applicant_name.like(like),
                Reimbursement.entity_name.like(like),
                Reimbursement.tax_number.like(like),
                Reimbursement.items.any(ReimbursementItem.related_number.like(like)),
            )
        )
    if start_date:
        statement = statement.where(func.date(Reimbursement.created_at) >= start_date)
    if end_date:
        statement = statement.where(func.date(Reimbursement.created_at) <= end_date)
    rows = db.scalars(statement.order_by(Reimbursement.created_at.desc())).unique().all()

    month_prefix = date.today().strftime("%Y-%m")
    approved_this_month = [
        row
        for row in all_rows
        if row.status == "approved"
        and bool(row.finance_approved_at or row.supervisor_approved_at)
        and (row.finance_approved_at or row.supervisor_approved_at).strftime("%Y-%m")
        == month_prefix
    ]
    summary = {
        "pending_supervisor": sum(row.status == "pending_supervisor" for row in all_rows),
        "pending_finance": sum(row.status == "pending_finance" for row in all_rows),
        "pending_export": sum(not row.exported for row in all_rows),
        "month_approved_count": len(approved_this_month),
        "month_approved_amount": round(
            sum(float(row.total_amount) for row in approved_this_month), 2
        ),
    }
    return {
        "records": [serialize_claim(db, row, user) for row in rows],
        "summary": summary,
        "config": reimbursement_config_payload(db),
        "permissions": {
            "can_configure": bool(get_permission_scope(db, user, "reimbursement.configure")),
            "can_export": bool(get_permission_scope(db, user, "reimbursement.export")),
        },
    }


@router.get("/config")
def get_config(
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("reimbursement.view")),
) -> dict[str, object]:
    return reimbursement_config_payload(db)


@router.put("/config")
def save_config(
    payload: ReimbursementConfigInput,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("reimbursement.configure")),
) -> dict[str, bool]:
    setting = db.get(SystemSetting, FINANCE_SETTING_KEY)
    if setting is None:
        setting = SystemSetting(
            key=FINANCE_SETTING_KEY,
            value="true" if payload.finance_approval_enabled else "false",
            description="报销是否需要财务审批",
        )
        db.add(setting)
    else:
        setting.value = "true" if payload.finance_approval_enabled else "false"
    db.commit()
    return {"finance_approval_enabled": payload.finance_approval_enabled}


@router.get("/entities")
def list_entities(
    include_inactive: bool = False,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("reimbursement.view")),
) -> list[dict[str, object]]:
    statement = select(ReimbursementEntity)
    if not include_inactive:
        statement = statement.where(ReimbursementEntity.is_active.is_(True))
    entities = db.scalars(
        statement.order_by(ReimbursementEntity.is_default.desc(), ReimbursementEntity.name)
    ).all()
    return [serialize_entity(entity) for entity in entities]


def save_entity_record(
    db: Session,
    payload: ReimbursementEntityInput,
    entity: ReimbursementEntity | None = None,
) -> ReimbursementEntity:
    duplicate = db.scalar(
        select(ReimbursementEntity).where(
            ReimbursementEntity.name == payload.name,
            ReimbursementEntity.id != (entity.id if entity else 0),
        )
    )
    if duplicate:
        raise HTTPException(status_code=409, detail="该报销主体已经存在")
    if payload.is_default:
        for current in db.scalars(select(ReimbursementEntity)).all():
            current.is_default = False
    if entity is None:
        entity = ReimbursementEntity()
        db.add(entity)
    entity.name = payload.name
    entity.tax_number = payload.tax_number
    entity.is_default = payload.is_default
    entity.is_active = payload.is_active
    db.commit()
    db.refresh(entity)
    return entity


@router.post("/entities", status_code=status.HTTP_201_CREATED)
def create_entity(
    payload: ReimbursementEntityInput,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("reimbursement.configure")),
) -> dict[str, object]:
    return serialize_entity(save_entity_record(db, payload))


@router.put("/entities/{entity_id}")
def update_entity(
    entity_id: int,
    payload: ReimbursementEntityInput,
    db: Session = Depends(get_db),
    _: User = Depends(require_permission("reimbursement.configure")),
) -> dict[str, object]:
    entity = db.get(ReimbursementEntity, entity_id)
    if entity is None:
        raise HTTPException(status_code=404, detail="报销主体不存在")
    return serialize_entity(save_entity_record(db, payload, entity))


@router.get("/batch/template/xlsx")
def download_batch_template(
    _: User = Depends(require_permission("reimbursement.create")),
) -> FileResponse:
    if not BATCH_TEMPLATE_PATH.is_file():
        raise HTTPException(status_code=500, detail="批量导入模板尚未生成")
    return FileResponse(
        BATCH_TEMPLATE_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="报销批量导入模板.xlsx",
    )


@router.post("/batch/preview")
def preview_batch_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("reimbursement.create")),
) -> dict[str, object]:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix != ".xlsx":
        raise HTTPException(status_code=422, detail="请上传 .xlsx 格式的批量导入表")
    contents = file.file.read(MAX_UPLOAD_BYTES + 1)
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="导入文件不能超过 10MB")
    groups, errors = parse_batch_workbook(contents)
    errors.extend(validate_batch_groups(db, groups, user))
    preview = batch_preview_payload(groups)
    return {
        "claims": preview,
        "errors": errors,
        "claim_count": len(preview),
        "item_count": sum(int(item["item_count"]) for item in preview),
        "total_amount": round(sum(float(item["total_amount"]) for item in preview), 2),
        "can_import": bool(preview) and not errors,
    }


@router.post("/batch/import", status_code=status.HTTP_201_CREATED)
def import_batch_reimbursements(
    file: UploadFile = File(...),
    submit: bool = Form(default=False),
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("reimbursement.create")),
) -> dict[str, object]:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix != ".xlsx":
        raise HTTPException(status_code=422, detail="请上传 .xlsx 格式的批量导入表")
    contents = file.file.read(MAX_UPLOAD_BYTES + 1)
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="导入文件不能超过 10MB")
    groups, errors = parse_batch_workbook(contents)
    errors.extend(validate_batch_groups(db, groups, user))
    if not groups:
        raise HTTPException(status_code=422, detail="模板中没有可导入的报销数据")
    if errors:
        first = errors[0]
        location = f"第 {first['row']} 行" if "row" in first else f"分组 {first['group']}"
        raise HTTPException(
            status_code=422,
            detail=f"{location}：{first['message']}，请修正后重新预览",
        )

    imported: list[Reimbursement] = []
    timestamp = now()
    finance_required = finance_approval_enabled(db) if submit else False
    try:
        for group in groups:
            items = group["items"]
            assert isinstance(items, list)
            claim = Reimbursement(
                number=generate_number(db),
                applicant_id=int(group["applicant_id"]),
                applicant_name=str(group["applicant_name"]),
                team=str(group["team"]),
                entity_name=str(group["entity_name"]),
                tax_number=str(group["tax_number"]),
                status="pending_supervisor" if submit else "draft",
                note=str(group["note"]) or None,
                total_amount=sum((item["amount"] for item in items), start=Decimal("0")).quantize(
                    Decimal("0.01")
                ),
                finance_approval_required=finance_required,
                submitted_at=timestamp if submit else None,
            )
            for index, item in enumerate(items):
                claim.items.append(
                    ReimbursementItem(
                        expense_date=item["expense_date"],
                        category=str(item["category"]),
                        amount=item["amount"],
                        related_number=str(item["related_number"]) or None,
                        description=str(item["description"]) or None,
                        sort_order=index,
                    )
                )
            db.add(claim)
            add_action(
                claim,
                user,
                "create",
                "",
                "draft",
                f"Excel 批量导入 · 分组 {group['group_key']}",
            )
            if submit:
                add_action(claim, user, "submit", "draft", "pending_supervisor", "批量导入后提交")
            db.flush()
            imported.append(claim)
        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "claim_count": len(imported),
        "item_count": sum(len(claim.items) for claim in imported),
        "total_amount": round(sum(float(claim.total_amount) for claim in imported), 2),
        "status": "pending_supervisor" if submit else "draft",
        "records": [
            {"id": claim.id, "number": claim.number, "applicant_name": claim.applicant_name}
            for claim in imported
        ],
    }


@router.get("/template/xlsx")
def download_template(
    _: User = Depends(require_permission("reimbursement.create")),
) -> StreamingResponse:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "报销明细导入"
    sheet.append(["费用日期", "费用类别", "金额", "关联单号", "费用说明"])
    sheet.append([date.today(), "临时运费", 0, "快递单号（可选）", "具体用途（可选）"])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:E2"
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1D5FD1")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    for index, width in enumerate((14, 20, 14, 24, 36), start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    sheet["A2"].number_format = "yyyy-mm-dd"
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename*=UTF-8''reimbursement_import_template.xlsx"
        },
    )


@router.post("/import/preview")
def preview_import(
    file: UploadFile = File(...),
    _: User = Depends(require_permission("reimbursement.create")),
) -> dict[str, object]:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix != ".xlsx":
        raise HTTPException(status_code=422, detail="请上传 .xlsx 格式的导入表")
    contents = file.file.read(MAX_UPLOAD_BYTES + 1)
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="导入文件不能超过 10MB")
    try:
        workbook = load_workbook(BytesIO(contents), data_only=True, read_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
    except Exception as exc:
        raise HTTPException(status_code=422, detail="无法读取该 Excel 文件") from exc
    aliases = {
        "费用日期": "expense_date",
        "日期": "expense_date",
        "费用类别": "category",
        "类别": "category",
        "金额": "amount",
        "关联单号": "related_number",
        "单号": "related_number",
        "费用说明": "description",
        "说明": "description",
        "备注": "description",
    }
    mapping = {index: aliases[name] for index, name in enumerate(headers) if name in aliases}
    if "category" not in mapping.values() or "amount" not in mapping.values():
        raise HTTPException(status_code=422, detail="表格必须包含“费用类别”和“金额”两列")
    parsed: list[dict[str, object]] = []
    errors: list[dict[str, object]] = []
    for excel_row, values in enumerate(rows, start=2):
        raw = {
            field: values[index] if index < len(values) else None
            for index, field in mapping.items()
        }
        if not any(value not in {None, ""} for value in raw.values()):
            continue
        try:
            category = str(raw.get("category") or "").strip()
            if not category:
                raise ValueError("费用类别不能为空")
            amount = Decimal(str(raw.get("amount") or "0")).quantize(Decimal("0.01"))
            if amount <= 0:
                raise ValueError("金额必须大于 0")
            parsed.append(
                {
                    "expense_date": parse_date(raw.get("expense_date")),
                    "category": category,
                    "amount": float(amount),
                    "related_number": str(raw.get("related_number") or "").strip(),
                    "description": str(raw.get("description") or "").strip(),
                }
            )
        except (ValueError, TypeError, ArithmeticError) as exc:
            errors.append({"row": excel_row, "message": str(exc)})
    return {"items": parsed, "errors": errors, "count": len(parsed)}


@router.get("/export/xlsx")
def export_reimbursements(
    ids: str = Query(default=""),
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("reimbursement.export")),
) -> StreamingResponse:
    statement = visible_statement(db, user)
    selected_ids = [int(value) for value in ids.split(",") if value.strip().isdigit()]
    if selected_ids:
        statement = statement.where(Reimbursement.id.in_(selected_ids))
    else:
        statement = statement.where(Reimbursement.exported.is_(False))
    claims = db.scalars(statement.order_by(Reimbursement.number)).unique().all()
    if not claims:
        raise HTTPException(status_code=422, detail="没有可导出的报销单")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "报销数据明细"
    headers = [
        "报销单ID",
        "明细ID",
        "报销单号",
        "报销人",
        "所属组",
        "主体",
        "税号",
        "状态代码",
        "单据状态",
        "费用日期",
        "费用类别",
        "金额",
        "关联单号",
        "费用说明",
        "整单备注",
        "主管审批人",
        "主管审批时间",
        "财务审批人",
        "财务审批时间",
        "凭证数",
        "提交时间",
        "创建时间",
        "发票张数",
        "发票识别金额",
        "发票金额差异",
        "发票识别问题数",
    ]
    sheet.append(headers)
    supervisor_names: dict[int, str] = {}
    finance_names: dict[int, str] = {}
    for claim in claims:
        for record in claim.approval_records:
            if record.action == "supervisor_approve":
                supervisor_names[claim.id] = record.actor_name
            if record.action == "finance_approve":
                finance_names[claim.id] = record.actor_name
        for item in claim.items:
            sheet.append(
                [
                    claim.id,
                    item.id,
                    claim.number,
                    claim.applicant_name,
                    claim.team,
                    claim.entity_name,
                    claim.tax_number,
                    claim.status,
                    display_status(claim),
                    item.expense_date,
                    item.category,
                    float(item.amount),
                    item.related_number or "",
                    item.description or "",
                    claim.note or "",
                    supervisor_names.get(claim.id, ""),
                    claim.supervisor_approved_at,
                    finance_names.get(claim.id, ""),
                    claim.finance_approved_at,
                    len(claim.attachments),
                    claim.submitted_at,
                    claim.created_at,
                    sum(item.invoice is not None for item in claim.attachments),
                    float(
                        sum(
                            (
                                item.invoice.final_amount
                                for item in claim.attachments
                                if item.invoice and item.invoice.final_amount is not None
                            ),
                            start=Decimal("0"),
                        )
                    ),
                    float(
                        claim.total_amount
                        - sum(
                            (
                                item.invoice.final_amount
                                for item in claim.attachments
                                if item.invoice and item.invoice.final_amount is not None
                            ),
                            start=Decimal("0"),
                        )
                    ),
                    sum(
                        bool(
                            item.invoice
                            and item.invoice.recognition_status
                            in {"failed", "unconfigured", "needs_review"}
                        )
                        for item in claim.attachments
                    ),
                ]
            )
    header_fill = PatternFill("solid", fgColor="183B69")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="DDE3EC")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = (
        12,
        12,
        20,
        12,
        12,
        30,
        24,
        20,
        14,
        13,
        20,
        13,
        22,
        32,
        28,
        14,
        20,
        14,
        20,
        10,
        20,
        20,
        12,
        16,
        16,
        14,
    )
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 10).number_format = "yyyy-mm-dd"
        sheet.cell(row, 12).number_format = "¥#,##0.00"
        for column in (17, 19, 21, 22):
            sheet.cell(row, column).number_format = "yyyy-mm-dd hh:mm"
        for column in (24, 25):
            sheet.cell(row, column).number_format = "¥#,##0.00"
    summary = workbook.create_sheet("导出汇总")
    summary.append(["报销单数", len(claims)])
    summary.append(["明细条数", sum(len(claim.items) for claim in claims)])
    summary.append(["合计金额", sum(float(claim.total_amount) for claim in claims)])
    summary.append(["导出时间", now()])
    summary.append(["包含状态", "、".join(sorted({display_status(claim) for claim in claims}))])
    summary["B3"].number_format = "¥#,##0.00"
    summary["B4"].number_format = "yyyy-mm-dd hh:mm"
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 24

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = quote(f"报销数据_{date.today():%Y%m%d}.xlsx")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.post("/export/mark")
def mark_exported(
    payload: ReimbursementExportInput,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("reimbursement.export")),
) -> dict[str, object]:
    claims = (
        db.scalars(visible_statement(db, user).where(Reimbursement.id.in_(payload.ids)))
        .unique()
        .all()
    )
    if not claims:
        raise HTTPException(status_code=422, detail="没有可标记的报销单")
    timestamp = now()
    batch = f"DC{timestamp:%Y%m%d%H%M%S}"
    for claim in claims:
        claim.exported = True
        claim.exported_at = timestamp
        claim.export_batch = batch
        add_action(claim, user, "export", claim.status, claim.status, f"导出批次 {batch}")
    db.commit()
    return {"batch": batch, "count": len(claims)}


@router.post("", status_code=status.HTTP_201_CREATED)
def create_reimbursement(
    payload: ReimbursementInput,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("reimbursement.create")),
) -> dict[str, object]:
    claim = Reimbursement(
        number=generate_number(db),
        applicant_id=user.id,
        applicant_name=payload.applicant_name,
        team=payload.team,
        status="draft",
    )
    replace_items(claim, payload)
    db.add(claim)
    add_action(claim, user, "create", "", "draft", "自动保存草稿")
    db.commit()
    return serialize_claim(db, get_visible_claim(db, claim.id, user), user, detail=True)


@router.get("/attachments/{attachment_id}")
def download_attachment(
    attachment_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("reimbursement.view")),
) -> FileResponse:
    attachment = db.get(ReimbursementAttachment, attachment_id)
    if attachment is None:
        raise HTTPException(status_code=404, detail="附件不存在")
    get_visible_claim(db, attachment.reimbursement_id, user)
    path = (settings.storage_path / attachment.relative_path).resolve()
    root = settings.storage_path.resolve()
    if root not in path.parents or not path.is_file():
        raise HTTPException(status_code=404, detail="附件文件不存在")
    return FileResponse(path, media_type=attachment.content_type, filename=attachment.original_name)


@router.get("/{claim_id}")
def get_reimbursement(
    claim_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("reimbursement.view")),
) -> dict[str, object]:
    return serialize_claim(db, get_visible_claim(db, claim_id, user), user, detail=True)


@router.put("/{claim_id}")
def update_reimbursement(
    claim_id: int,
    payload: ReimbursementInput,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("reimbursement.create")),
) -> dict[str, object]:
    claim = get_visible_claim(db, claim_id, user)
    if not can_edit(db, claim, user):
        raise HTTPException(status_code=409, detail="当前状态不能修改")
    replace_items(claim, payload)
    invalidate_export(claim)
    db.commit()
    return serialize_claim(db, get_visible_claim(db, claim.id, user), user, detail=True)


@router.delete("/{claim_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_reimbursement(
    claim_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("reimbursement.create")),
) -> None:
    claim = get_visible_claim(db, claim_id, user)
    if claim.status != "draft" or not can_edit(db, claim, user):
        raise HTTPException(status_code=409, detail="当前报销单不能删除")
    paths = [(settings.storage_path / item.relative_path).resolve() for item in claim.attachments]
    db.execute(delete(Reimbursement).where(Reimbursement.id == claim.id))
    db.commit()
    root = settings.storage_path.resolve()
    for path in paths:
        if root in path.parents and path.is_file():
            path.unlink()


@router.post("/{claim_id}/submit")
def submit_reimbursement(
    claim_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("reimbursement.create")),
) -> dict[str, object]:
    claim = get_visible_claim(db, claim_id, user)
    if not can_edit(db, claim, user):
        raise HTTPException(status_code=409, detail="当前状态不能提交")
    validate_for_submit(claim)
    old_status = claim.status
    claim.status = "pending_supervisor"
    claim.finance_approval_required = finance_approval_enabled(db)
    claim.submitted_at = now()
    invalidate_export(claim)
    add_action(claim, user, "submit", old_status, claim.status)
    db.commit()
    return serialize_claim(db, get_visible_claim(db, claim.id, user), user, detail=True)


@router.post("/{claim_id}/approve")
def approve_reimbursement(
    claim_id: int,
    payload: ReimbursementActionInput,
    db: Session = Depends(get_db),
    user: User = Depends(
        require_any_permission("reimbursement.approve_supervisor", "reimbursement.approve_finance")
    ),
) -> dict[str, object]:
    claim = get_visible_claim(db, claim_id, user)
    if not can_approve(db, claim, user):
        raise HTTPException(status_code=403, detail="当前用户不能审批这张报销单")
    old_status = claim.status
    timestamp = now()
    if old_status == "pending_supervisor":
        claim.supervisor_approved_at = timestamp
        claim.status = "pending_finance" if claim.finance_approval_required else "approved"
        action = "supervisor_approve"
    else:
        claim.finance_approved_at = timestamp
        claim.status = "approved"
        action = "finance_approve"
    invalidate_export(claim)
    add_action(claim, user, action, old_status, claim.status, payload.comment)
    db.commit()
    return serialize_claim(db, get_visible_claim(db, claim.id, user), user, detail=True)


@router.post("/{claim_id}/return")
def return_reimbursement(
    claim_id: int,
    payload: ReimbursementActionInput,
    db: Session = Depends(get_db),
    user: User = Depends(
        require_any_permission("reimbursement.approve_supervisor", "reimbursement.approve_finance")
    ),
) -> dict[str, object]:
    claim = get_visible_claim(db, claim_id, user)
    if not can_approve(db, claim, user):
        raise HTTPException(status_code=403, detail="当前用户不能退回这张报销单")
    if not payload.comment:
        raise HTTPException(status_code=422, detail="退回时必须填写原因")
    old_status = claim.status
    claim.status = "returned"
    invalidate_export(claim)
    add_action(claim, user, "return", old_status, claim.status, payload.comment)
    db.commit()
    return serialize_claim(db, get_visible_claim(db, claim.id, user), user, detail=True)


def recognize_attachment_invoice(
    db: Session,
    claim: Reimbursement,
    attachment: ReimbursementAttachment,
    contents: bytes,
    suffix: str,
) -> ReimbursementInvoice:
    invoice = attachment.invoice or ReimbursementInvoice(attachment=attachment)
    invoice.recognition_status = "pending"
    invoice.recognition_provider = settings.invoice_ocr_provider
    invoice.recognition_message = "正在识别"
    try:
        result = recognize_invoice(contents, suffix)
        invoice.recognition_status = result.status
        invoice.recognition_provider = result.provider
        invoice.recognition_message = result.message
        invoice.recognized_entity_name = result.entity_name or None
        invoice.recognized_tax_number = result.tax_number or None
        invoice.recognized_amount = result.amount
        invoice.final_entity_name = result.entity_name or None
        invoice.final_tax_number = result.tax_number or None
        invoice.final_amount = result.amount
        invoice.invoice_code = result.invoice_code or None
        invoice.invoice_number = result.invoice_number or None
        invoice.invoice_date = result.invoice_date.date() if result.invoice_date else None
        invoice.provider_request_id = result.request_id or None
        invoice.recognized_at = now()
        invoice.manually_edited = False

        if invoice.invoice_number:
            duplicate = db.scalar(
                select(ReimbursementInvoice.id)
                .join(ReimbursementAttachment)
                .where(
                    ReimbursementInvoice.invoice_number == invoice.invoice_number,
                    ReimbursementInvoice.id != (invoice.id or 0),
                    ReimbursementAttachment.reimbursement_id != claim.id,
                )
                .limit(1)
            )
            if duplicate:
                invoice.recognition_status = "needs_review"
                invoice.recognition_message = "检测到相同发票号码，请核对是否重复报销"
        if not claim.entity_name and invoice.final_entity_name:
            claim.entity_name = invoice.final_entity_name
        if not claim.tax_number and invoice.final_tax_number:
            claim.tax_number = invoice.final_tax_number
    except InvoiceOcrNotConfiguredError as exc:
        invoice.recognition_status = "unconfigured"
        invoice.recognition_message = str(exc)
        invoice.recognized_at = now()
    except InvoiceOcrError as exc:
        invoice.recognition_status = "failed"
        invoice.recognition_message = str(exc)
        invoice.recognized_at = now()
    db.add(invoice)
    return invoice


@router.post("/{claim_id}/attachments", status_code=status.HTTP_201_CREATED)
def upload_attachment(
    claim_id: int,
    file: UploadFile = File(...),
    document_type: str = Form(default="voucher"),
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("reimbursement.create")),
) -> dict[str, object]:
    claim = get_visible_claim(db, claim_id, user)
    if not can_edit(db, claim, user):
        raise HTTPException(status_code=409, detail="当前状态不能上传附件")
    original_name = Path(file.filename or "附件").name
    suffix = Path(original_name).suffix.lower()
    if suffix not in ALLOWED_UPLOAD_SUFFIXES:
        raise HTTPException(status_code=422, detail="仅支持图片、PDF 和 XLSX 附件")
    contents = file.file.read(MAX_UPLOAD_BYTES + 1)
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="单个附件不能超过 10MB")
    if not contents:
        raise HTTPException(status_code=422, detail="附件内容为空")
    if document_type not in {"invoice", "voucher"}:
        raise HTTPException(status_code=422, detail="附件类型无效")
    if document_type == "invoice" and suffix not in {".jpg", ".jpeg", ".png", ".webp", ".pdf"}:
        raise HTTPException(status_code=422, detail="发票识别仅支持图片和 PDF")
    digest = sha256(contents).hexdigest()
    duplicate = db.scalar(
        select(ReimbursementAttachment).where(
            ReimbursementAttachment.sha256 == digest,
            ReimbursementAttachment.reimbursement_id != claim.id,
        )
    )
    folder = Path("reimbursements") / f"{date.today():%Y}" / claim.number
    stored_name = f"{uuid4().hex}{suffix}"
    relative_path = folder / stored_name
    target = (settings.storage_path / relative_path).resolve()
    root = settings.storage_path.resolve()
    if root not in target.parents:
        raise HTTPException(status_code=400, detail="附件路径无效")
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(contents)
    attachment = ReimbursementAttachment(
        reimbursement_id=claim.id,
        uploaded_by=user.id,
        original_name=original_name,
        stored_name=stored_name,
        relative_path=str(relative_path).replace("\\", "/"),
        content_type=file.content_type or "application/octet-stream",
        size_bytes=len(contents),
        sha256=digest,
        document_type=document_type,
    )
    db.add(attachment)
    db.flush()
    attachment_id = attachment.id
    if document_type == "invoice":
        recognize_attachment_invoice(db, claim, attachment, contents, suffix)
    invalidate_export(claim)
    db.commit()
    db.expire_all()
    refreshed = get_visible_claim(db, claim.id, user)
    uploaded = next(item for item in refreshed.attachments if item.id == attachment_id)
    return serialize_attachment(uploaded, duplicate=bool(duplicate))


@router.post("/{claim_id}/invoices/{invoice_id}/recognize")
def retry_invoice_recognition(
    claim_id: int,
    invoice_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("reimbursement.create")),
) -> dict[str, object]:
    claim = get_visible_claim(db, claim_id, user)
    if not can_edit(db, claim, user):
        raise HTTPException(status_code=409, detail="当前状态不能重新识别发票")
    attachment = next(
        (item for item in claim.attachments if item.invoice and item.invoice.id == invoice_id),
        None,
    )
    if attachment is None:
        raise HTTPException(status_code=404, detail="发票不存在")
    path = (settings.storage_path / attachment.relative_path).resolve()
    root = settings.storage_path.resolve()
    if root not in path.parents or not path.is_file():
        raise HTTPException(status_code=404, detail="发票文件不存在")
    recognize_attachment_invoice(db, claim, attachment, path.read_bytes(), path.suffix.lower())
    invalidate_export(claim)
    db.commit()
    refreshed = get_visible_claim(db, claim.id, user)
    result = next(item for item in refreshed.attachments if item.id == attachment.id)
    return serialize_attachment(result)


@router.put("/{claim_id}/invoices/{invoice_id}")
def update_invoice_recognition(
    claim_id: int,
    invoice_id: int,
    payload: ReimbursementInvoiceInput,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("reimbursement.create")),
) -> dict[str, object]:
    claim = get_visible_claim(db, claim_id, user)
    if not can_edit(db, claim, user):
        raise HTTPException(status_code=409, detail="当前状态不能修改发票识别结果")
    attachment = next(
        (item for item in claim.attachments if item.invoice and item.invoice.id == invoice_id),
        None,
    )
    if attachment is None or attachment.invoice is None:
        raise HTTPException(status_code=404, detail="发票不存在")
    invoice = attachment.invoice
    invoice.final_entity_name = payload.entity_name or None
    invoice.final_tax_number = payload.tax_number or None
    invoice.final_amount = payload.amount
    invoice.manually_edited = bool(
        (invoice.recognized_entity_name or "") != payload.entity_name
        or (invoice.recognized_tax_number or "") != payload.tax_number
        or invoice.recognized_amount != payload.amount
    )
    if payload.entity_name and payload.tax_number and payload.amount is not None:
        invoice.recognition_status = "confirmed" if invoice.manually_edited else "success"
        invoice.recognition_message = (
            "已人工修改并确认" if invoice.manually_edited else "识别结果已确认"
        )
    else:
        invoice.recognition_status = "needs_review"
        invoice.recognition_message = "主体、税号和金额需要补充完整"
    if not claim.entity_name and payload.entity_name:
        claim.entity_name = payload.entity_name
    if not claim.tax_number and payload.tax_number:
        claim.tax_number = payload.tax_number
    invalidate_export(claim)
    db.commit()
    refreshed = get_visible_claim(db, claim.id, user)
    result = next(item for item in refreshed.attachments if item.id == attachment.id)
    return serialize_attachment(result)


@router.delete("/{claim_id}/attachments/{attachment_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_attachment(
    claim_id: int,
    attachment_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_permission("reimbursement.create")),
) -> None:
    claim = get_visible_claim(db, claim_id, user)
    if not can_edit(db, claim, user):
        raise HTTPException(status_code=409, detail="当前状态不能删除附件")
    attachment = db.get(ReimbursementAttachment, attachment_id)
    if attachment is None or attachment.reimbursement_id != claim.id:
        raise HTTPException(status_code=404, detail="附件不存在")
    path = (settings.storage_path / attachment.relative_path).resolve()
    db.delete(attachment)
    invalidate_export(claim)
    db.commit()
    root = settings.storage_path.resolve()
    if root in path.parents and path.is_file():
        path.unlink()
