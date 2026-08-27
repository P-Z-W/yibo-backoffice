"""Employee salary CRUD and Excel export APIs."""

from __future__ import annotations

from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.operations import SalaryRecord
from app.models.user import User
from app.schemas.operations import SalaryInput

router = APIRouter(prefix="/salary", tags=["员工工资"])


def serialize(row: SalaryRecord) -> dict[str, object]:
    base = float(row.base_salary)
    bonus = float(row.bonus)
    deduction = float(row.deduction)
    return {
        "id": row.id,
        "name": row.name,
        "team": row.team,
        "year_month": row.year_month,
        "base_salary": base,
        "bonus": bonus,
        "deduction": deduction,
        "total": round(base + bonus - deduction, 2),
        "note": row.note or "",
        "created_at": row.created_at.isoformat() if row.created_at else "",
    }


@router.get("")
def list_salary(
    month: str = "",
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> dict[str, object]:
    statement = select(SalaryRecord)
    if month:
        statement = statement.where(SalaryRecord.year_month == month)
    rows = db.scalars(
        statement.order_by(SalaryRecord.year_month.desc(), SalaryRecord.team, SalaryRecord.name)
    ).all()
    records = [serialize(row) for row in rows]
    return {
        "records": records,
        "summary": {
            "employees": len(records),
            "base_salary": round(sum(float(row["base_salary"]) for row in records), 2),
            "bonus": round(sum(float(row["bonus"]) for row in records), 2),
            "deduction": round(sum(float(row["deduction"]) for row in records), 2),
            "total": round(sum(float(row["total"]) for row in records), 2),
        },
    }


@router.post("", status_code=status.HTTP_201_CREATED)
def add_salary(
    payload: SalaryInput,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> dict[str, object]:
    target = SalaryRecord(**payload.model_dump())
    target.note = payload.note.strip() or None
    db.add(target)
    db.commit()
    db.refresh(target)
    return serialize(target)


@router.put("/{record_id}")
def save_salary(
    record_id: int,
    payload: SalaryInput,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> dict[str, object]:
    target = db.get(SalaryRecord, record_id)
    if target is None:
        raise HTTPException(status_code=404, detail="工资记录不存在")
    for field, value in payload.model_dump().items():
        setattr(target, field, value)
    target.note = payload.note.strip() or None
    db.commit()
    return serialize(target)


@router.delete("/{record_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_salary(
    record_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> None:
    if db.get(SalaryRecord, record_id) is None:
        raise HTTPException(status_code=404, detail="工资记录不存在")
    db.execute(delete(SalaryRecord).where(SalaryRecord.id == record_id))
    db.commit()


@router.get("/export/xlsx")
def export_salary(
    month: str = "",
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StreamingResponse:
    statement = select(SalaryRecord)
    if month:
        statement = statement.where(SalaryRecord.year_month == month)
    rows = db.scalars(statement.order_by(SalaryRecord.team, SalaryRecord.name)).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"工资表{month}"
    headers = ["姓名", "团队", "工资月份", "基本工资", "绩效奖金", "扣款", "实发工资", "备注"]
    sheet.append(headers)
    fill = PatternFill("solid", fgColor="1D3152")
    font = Font(bold=True, color="FFFFFF")
    side = Side(style="thin", color="DDDDDD")
    border = Border(left=side, right=side, top=side, bottom=side)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for row in rows:
        base = float(row.base_salary)
        bonus = float(row.bonus)
        deduction = float(row.deduction)
        sheet.append(
            [
                row.name,
                row.team,
                row.year_month,
                base,
                bonus,
                deduction,
                round(base + bonus - deduction, 2),
                row.note or "",
            ]
        )
    for index, width in enumerate((12, 14, 12, 12, 12, 10, 12, 24), start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = quote(f"员工工资表_{month or '全部'}.xlsx")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )
