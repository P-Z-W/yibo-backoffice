"""Excel parsing and summary helpers for operating-analysis detail tables."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 5000
MAX_IMPORT_COLUMNS = 80
PREVIEW_ROWS = 20

STAFFING_INPUT_COLUMNS: tuple[str, ...] = (
    "小组",
    "正式工人数",
    "最优配置",
    "人均月产出",
    "最优人均产出",
    "综合分析",
)
STAFFING_CALCULATED_COLUMNS: tuple[str, ...] = (
    "配置偏差",
    "偏差比例",
    "效率差额",
    "效率差额占比",
    "人均月产出净变化",
    "人均月产出环比",
)
STAFFING_COLUMNS: tuple[str, ...] = (
    "小组",
    "正式工人数",
    "最优配置",
    "配置偏差",
    "偏差比例",
    "人均月产出",
    "最优人均产出",
    "效率差额",
    "效率差额占比",
    "人均月产出净变化",
    "人均月产出环比",
    "综合分析",
)
STAFFING_PERCENT_COLUMNS = {
    "偏差比例",
    "效率差额占比",
    "人均月产出环比",
}


@dataclass(frozen=True)
class DatasetDefinition:
    code: str
    name: str
    description: str
    sheet_aliases: tuple[str, ...]
    columns: tuple[str, ...]
    required_columns: tuple[str, ...]
    summary_hint: str
    metric_codes: tuple[str, ...]


DATASET_DEFINITIONS: tuple[DatasetDefinition, ...] = (
    DatasetDefinition(
        "shipping_orders",
        "发货单量",
        "按页面月份从业务库只读统计各团队发货单量，也可上传 Excel 补充。",
        ("1-发货单量", "发货单量"),
        ("团队名称", "发货单量", "数据发货占比", "备注"),
        ("团队名称", "发货单量"),
        "系统按运单号去重，并按已确认的状态、物流和仓库条件统计。",
        ("shipping_orders",),
    ),
    DatasetDefinition(
        "return_items",
        "退货件数",
        "按页面月份从业务库只读统计各团队退货件数，也可上传 Excel 补充。",
        ("2-退货件数", "退货件数"),
        (
            "团队名称",
            "处理退货件数",
            "拦截件扣费件数",
            "异常件扣费件数",
            "退货件数合计",
            "数据退货占比",
        ),
        ("团队名称", "处理退货件数", "拦截件扣费件数", "异常件扣费件数"),
        "三类退货件数按团队汇总；退货件数合计及数据退货占比由系统计算。",
        ("return_items",),
    ),
    DatasetDefinition(
        "customer_changes",
        "客户变化",
        "数据来自客户管理模块，按页面月份展示新进、流失和意向客户。",
        ("3-客户变化", "客户变化"),
        ("月份", "变化类型", "客户名称", "来源渠道", "数量", "备注"),
        ("变化类型",),
        "按客户管理模块的变化类型汇总；数量为空时按 1 计。",
        ("new_customers", "lost_customers", "prospective_customers"),
    ),
    DatasetDefinition(
        "supplier_changes",
        "供应商变化",
        "数据来自供应商管理模块，按页面月份展示当月发生变化的供应商。",
        ("4-供应商变化", "供应商变化"),
        (
            "供应商名称",
            "供应商联系人",
            "联系电话",
            "联系地址",
            "合作时间",
            "常用产品类型",
            "备注",
        ),
        ("供应商名称",),
        "按供应商管理模块当月变更记录去重后汇总供应商数量。",
        ("supplier_change",),
    ),
    DatasetDefinition(
        "staffing",
        "人员调整",
        "按页面月份上传各小组人员配置和产出，偏差、效率差额及环比由系统计算。",
        ("5-人员调整", "人员调整"),
        STAFFING_COLUMNS,
        ("小组", "正式工人数"),
        "页面月份为准；识别正式工人数后汇总，其他偏差和环比由系统统一计算。",
        ("staff_adjustment",),
    ),
    DatasetDefinition(
        "value_added",
        "增值服务细项",
        "按团队和服务类型登记增值服务数量。",
        ("6-增值服务细项", "增值服务细项"),
        ("月份", "团队ID", "团队名称", "服务编码", "服务名称", "服务分组", "数量"),
        ("团队名称", "服务名称", "数量"),
        "当前只进入分表；金额与次数口径确认后再自动汇总。",
        (),
    ),
    DatasetDefinition(
        "service_issues",
        "客户服务情况",
        "数据来自客户服务管理模块，展示投诉、原因、责任和整改状态。",
        ("8-客户服务情况", "客户服务情况"),
        (
            "月份",
            "团队名",
            "投诉大类",
            "问题详细描述",
            "核实原因",
            "责任归属",
            "整改措施",
            "状态",
        ),
        ("月份", "投诉大类", "问题详细描述"),
        "按客户服务管理模块当月记录展示，不直接改变总表数值。",
        (),
    ),
    DatasetDefinition(
        "short_video",
        "短视频情况",
        "数据来自短视频管理模块，展示数量、类型、负责人和运营备注。",
        ("9-短视频情况", "短视频情况"),
        ("月份", "短视频数量", "短视频类型", "负责人", "备注"),
        ("月份", "短视频数量"),
        "按短视频管理模块当月记录展示，不直接改变现有总表数值。",
        (),
    ),
)

DEFINITIONS_BY_CODE = {item.code: item for item in DATASET_DEFINITIONS}


def public_definitions() -> list[dict[str, object]]:
    return [
        {
            "code": item.code,
            "name": item.name,
            "description": item.description,
            "columns": list(item.columns),
            "summary_hint": item.summary_hint,
            "metric_codes": list(item.metric_codes),
        }
        for item in DATASET_DEFINITIONS
    ]


def get_definition(code: str) -> DatasetDefinition:
    try:
        return DEFINITIONS_BY_CODE[code]
    except KeyError as exc:
        raise ValueError("不支持的经营分析分表") from exc


def normalized_header(value: object) -> str:
    return re.sub(r"[\s_\-—（）()%％/]+", "", str(value or "").strip()).lower()


def json_value(value: object) -> object:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        return value.strip()
    return value


def unique_headers(values: list[object]) -> tuple[list[str], list[bool]]:
    headers: list[str] = []
    generated: list[bool] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = str(value or "").strip()
        is_generated = not base
        if not base:
            base = f"未命名列{index}"
        counts[base] = counts.get(base, 0) + 1
        headers.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
        generated.append(is_generated)
    return headers, generated


def choose_sheet(workbook: object, definition: DatasetDefinition):
    worksheets = workbook.worksheets
    names = {normalized_header(sheet.title): sheet for sheet in worksheets}
    for alias in definition.sheet_aliases:
        match = names.get(normalized_header(alias))
        if match is not None:
            return match
    return workbook.active


def parse_workbook(
    filename: str,
    contents: bytes,
    definition: DatasetDefinition,
) -> dict[str, object]:
    if Path(filename).suffix.lower() != ".xlsx":
        raise ValueError("请上传 .xlsx 格式的文件")
    if len(contents) > MAX_UPLOAD_BYTES:
        raise ValueError("导入文件不能超过 10MB")
    try:
        workbook = load_workbook(BytesIO(contents), data_only=True, read_only=True)
        sheet = choose_sheet(workbook, definition)
        raw_rows = list(sheet.iter_rows(values_only=True, max_row=MAX_IMPORT_ROWS + 21))
    except Exception as exc:
        raise ValueError("无法读取该 Excel 文件") from exc
    if not raw_rows:
        raise ValueError("Excel 中没有可导入的数据")

    header_index: int | None = None
    expected = {normalized_header(value) for value in definition.columns}
    fallback: int | None = None
    best_score = -1
    for index, row in enumerate(raw_rows[:20]):
        values = [value for value in row if value not in (None, "")]
        if len(values) < 2:
            continue
        if fallback is None:
            fallback = index
        score = sum(normalized_header(value) in expected for value in values)
        if score > best_score:
            header_index = index
            best_score = score
    if best_score <= 0:
        header_index = fallback
    if header_index is None:
        raise ValueError("没有找到有效的表头行")

    raw_headers = list(raw_rows[header_index])
    last_header = max(
        (index for index, value in enumerate(raw_headers) if value not in (None, "")),
        default=-1,
    )
    if last_header < 0:
        raise ValueError("表头不能为空")
    if last_header + 1 > MAX_IMPORT_COLUMNS:
        raise ValueError(f"分表最多支持 {MAX_IMPORT_COLUMNS} 列")
    headers, generated = unique_headers(raw_headers[: last_header + 1])

    matrix: list[list[object]] = []
    for raw in raw_rows[header_index + 1 :]:
        values = [json_value(value) for value in list(raw)[: len(headers)]]
        values.extend([None] * (len(headers) - len(values)))
        if not any(value not in (None, "") for value in values):
            continue
        matrix.append(values)
        if len(matrix) > MAX_IMPORT_ROWS:
            raise ValueError(f"单次最多导入 {MAX_IMPORT_ROWS} 行")
    if not matrix:
        raise ValueError("表头下方没有可导入的数据")

    keep_indexes = [
        index
        for index in range(len(headers))
        if not generated[index]
        or any(row[index] not in (None, "") for row in matrix)
    ]
    columns = [headers[index] for index in keep_indexes]
    rows = [
        {columns[position]: row[index] for position, index in enumerate(keep_indexes)}
        for row in matrix
    ]
    present = {normalized_header(column) for column in columns}
    missing = [
        column for column in definition.required_columns if normalized_header(column) not in present
    ]
    warnings = []
    if missing:
        warnings.append(f"未识别推荐字段：{'、'.join(missing)}；数据仍可导入，但可能无法自动汇总。")
    return {
        "sheet_name": sheet.title,
        "columns": columns,
        "rows": rows,
        "row_count": len(rows),
        "preview_rows": rows[:PREVIEW_ROWS],
        "warnings": warnings,
    }


def find_column(rows: list[dict[str, object]], aliases: tuple[str, ...]) -> str | None:
    if not rows:
        return None
    columns = list(rows[0])
    normalized = {normalized_header(column): column for column in columns}
    for alias in aliases:
        match = normalized.get(normalized_header(alias))
        if match:
            return match
    return None


def number(value: object) -> Decimal | None:
    if value in (None, "", "-", "--"):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "").replace("，", "")
    if text.endswith(("%", "％")):
        text = text[:-1]
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def sum_column(rows: list[dict[str, object]], aliases: tuple[str, ...]) -> Decimal | None:
    column = find_column(rows, aliases)
    if column is None:
        return None
    values = [number(row.get(column)) for row in rows]
    usable = [value for value in values if value is not None]
    return sum(usable, Decimal("0")) if usable else None


def summarize_rows(dataset_type: str, rows: list[dict[str, object]]) -> dict[str, Decimal]:
    if dataset_type == "shipping_orders":
        value = sum_column(rows, ("发货单量", "单量"))
        return {"shipping_orders": value} if value is not None else {}
    if dataset_type == "return_items":
        value = sum_column(rows, ("退货件数合计", "退货件数", "退件件数"))
        if value is None:
            component_columns = [
                find_column(rows, (label,))
                for label in ("处理退货件数", "拦截件扣费件数", "异常件扣费件数")
            ]
            if any(component_columns):
                value = sum(
                    (
                        number(row.get(column)) or Decimal("0")
                        for row in rows
                        for column in component_columns
                        if column is not None
                    ),
                    Decimal("0"),
                )
        return {"return_items": value} if value is not None else {}
    if dataset_type == "staffing":
        value = sum_column(rows, ("正式工人数",))
        return {"staff_adjustment": value} if value is not None else {}
    if dataset_type == "supplier_changes":
        column = find_column(rows, ("供应商名称",))
        if column is None:
            return {}
        names = {str(row.get(column) or "").strip() for row in rows}
        names.discard("")
        return {"supplier_change": Decimal(len(names))} if names else {}
    if dataset_type != "customer_changes":
        return {}

    type_column = find_column(rows, ("变化类型", "客户类型"))
    quantity_column = find_column(rows, ("数量", "客户数量"))
    if type_column is None:
        return {}
    totals = {
        "new_customers": Decimal("0"),
        "lost_customers": Decimal("0"),
        "prospective_customers": Decimal("0"),
    }
    matched = False
    for row in rows:
        change_type = str(row.get(type_column) or "").strip()
        code = None
        if "新" in change_type:
            code = "new_customers"
        elif "流失" in change_type:
            code = "lost_customers"
        elif "意向" in change_type:
            code = "prospective_customers"
        if code is None:
            continue
        quantity = number(row.get(quantity_column)) if quantity_column else Decimal("1")
        totals[code] += quantity if quantity is not None else Decimal("1")
        matched = True
    return totals if matched else {}


def build_template(definition: DatasetDefinition) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = definition.name[:31]
    sheet.append(list(definition.columns))
    sheet.freeze_panes = "A2"
    fill = PatternFill("solid", fgColor="DCEAFF")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="24436C")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, column in enumerate(definition.columns, start=1):
        sheet.column_dimensions[chr(64 + index) if index <= 26 else "A"].width = max(
            14, min(28, len(column) * 2 + 4)
        )
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_staffing_template() -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "人员调整"
    columns = list(STAFFING_INPUT_COLUMNS)
    sheet.append(columns)
    sheet.append(["发货组", None, None, None, None, ""])
    sheet.append(["售后组", None, None, None, None, ""])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:F3"
    sheet.sheet_view.showGridLines = False
    widths = [20, 15, 15, 18, 18, 56]
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    input_fill = PatternFill("solid", fgColor="FFF9E8")
    thin_border = Border(
        left=Side(style="thin", color="DCE6F5"),
        right=Side(style="thin", color="DCE6F5"),
        top=Side(style="thin", color="DCE6F5"),
        bottom=Side(style="thin", color="DCE6F5"),
    )
    for column_index, (column, width) in enumerate(zip(columns, widths), start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.font = Font(bold=True, color="263F63")
        cell.fill = header_fill
        cell.border = Border(bottom=Side(style="medium", color="4D8DF7"))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    for row_index in range(2, 4):
        sheet.row_dimensions[row_index].height = 28
        for column_index in range(1, len(columns) + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.fill = input_fill
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="left" if column_index in {1, 6} else "right",
                vertical="center",
                wrap_text=column_index == 6,
            )
        for column_index in {2, 3}:
            sheet.cell(row=row_index, column=column_index).number_format = "#,##0.##"
        for column_index in {4, 5}:
            sheet.cell(row=row_index, column=column_index).number_format = "#,##0.00"
    sheet["A1"].comment = Comment("系统按页面所选月份和小组名称匹配；模板中不再填写月份。", "系统")
    sheet["B1"].comment = Comment("正式工人数用于月度总表汇总和配置偏差计算。", "系统")
    sheet["D1"].comment = Comment("人均月产出按月上传表内数据，后续再维护自动取数。", "系统")
    sheet["F1"].comment = Comment("支持填写较长的月度综合分析。", "系统")
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_detail_export(
    definition: DatasetDefinition, rows: list[dict[str, object]]
) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = definition.name[:31]
    columns = list(definition.columns)
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column) for column in columns])

    last_row = len(rows) + 1
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{last_row}"
    sheet.row_dimensions[1].height = 28
    sheet.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    header_border = Border(bottom=Side(style="medium", color="4D8DF7"))
    thin_border = Border(
        left=Side(style="thin", color="DCE6F5"),
        right=Side(style="thin", color="DCE6F5"),
        top=Side(style="thin", color="DCE6F5"),
        bottom=Side(style="thin", color="DCE6F5"),
    )
    for column_index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.font = Font(bold=True, color="263F63")
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(column_index)].width = max(
            14, min(32, len(column) * 2 + 6)
        )
    for row_index in range(2, last_row + 1):
        sheet.row_dimensions[row_index].height = 24
        for column_index, column in enumerate(columns, start=1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=column == "综合分析",
            )
            if column in STAFFING_PERCENT_COLUMNS:
                cell.number_format = "0.00%"
            elif column in {"正式工人数", "最优配置", "配置偏差"}:
                cell.number_format = "#,##0.##"
            elif column in {
                "人均月产出",
                "最优人均产出",
                "效率差额",
                "人均月产出净变化",
            }:
                cell.number_format = "#,##0.00"
        if "综合分析" in columns:
            sheet.row_dimensions[row_index].height = 48

    if "综合分析" in columns:
        sheet.column_dimensions[get_column_letter(columns.index("综合分析") + 1)].width = 56

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_shipping_template(rows: list[dict[str, object]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "发货单量"
    columns = ["序号", "团队名称", "发货单量", "数据发货占比", "备注"]
    sheet.append(columns)

    source_rows = rows or [{"团队名称": "", "发货单量": None, "备注": ""}]
    last_row = len(source_rows) + 1
    for index, row in enumerate(source_rows, start=1):
        excel_row = index + 1
        sheet.append(
            [
                index,
                str(row.get("团队名称") or ""),
                number(row.get("发货单量")),
                f'=IFERROR(C{excel_row}/SUM($C$2:$C${last_row}),0)',
                str(row.get("备注") or ""),
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:E{last_row}"
    sheet.row_dimensions[1].height = 28
    sheet.sheet_view.showGridLines = False
    widths = [9, 28, 16, 18, 42]
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    header_border = Border(bottom=Side(style="medium", color="4D8DF7"))
    thin_border = Border(
        left=Side(style="thin", color="DCE6F5"),
        right=Side(style="thin", color="DCE6F5"),
        top=Side(style="thin", color="DCE6F5"),
        bottom=Side(style="thin", color="DCE6F5"),
    )
    for column_index, (column, width) in enumerate(zip(columns, widths), start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.font = Font(bold=True, color="263F63")
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    for row_index in range(2, last_row + 1):
        sheet.row_dimensions[row_index].height = 24
        for column_index in range(1, 6):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="left" if column_index in {2, 5} else "center",
                vertical="center",
            )
        sheet.cell(row=row_index, column=3).number_format = "#,##0"
        sheet.cell(row=row_index, column=4).number_format = "0.00%"

    sheet["A1"].comment = Comment("序号仅用于展示，上传时不参与匹配。", "系统")
    sheet["B1"].comment = Comment("系统按团队名称匹配现有发货明细。", "系统")
    sheet["D1"].comment = Comment("该列由发货单量自动计算，上传时系统会重新计算。", "系统")
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_shipping_export(
    rows: list[dict[str, object]], columns: list[str]
) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "发货数据"
    export_columns = ["序号", *columns]
    sheet.append(export_columns)
    for index, row in enumerate(rows, start=1):
        values: list[object] = [index]
        for column in columns:
            value = row.get(column)
            if column == "数据发货占比" and value not in (None, ""):
                value = Decimal(str(value)) / Decimal("100")
            values.append(value)
        sheet.append(values)

    last_row = len(rows) + 1
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(export_columns))}{last_row}"
    sheet.row_dimensions[1].height = 28
    sheet.sheet_view.showGridLines = False
    width_map = {
        "序号": 9,
        "团队名称": 28,
        "发货单量": 16,
        "数据发货占比": 18,
        "备注": 42,
    }
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    header_border = Border(bottom=Side(style="medium", color="4D8DF7"))
    thin_border = Border(
        left=Side(style="thin", color="DCE6F5"),
        right=Side(style="thin", color="DCE6F5"),
        top=Side(style="thin", color="DCE6F5"),
        bottom=Side(style="thin", color="DCE6F5"),
    )
    for column_index, column in enumerate(export_columns, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.font = Font(bold=True, color="263F63")
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(column_index)].width = width_map[column]
    for row_index in range(2, last_row + 1):
        sheet.row_dimensions[row_index].height = 24
        for column_index, column in enumerate(export_columns, start=1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal=(
                    "center"
                    if column == "序号"
                    else "left"
                    if column in {"团队名称", "备注"}
                    else "right"
                ),
                vertical="center",
            )
            if column == "发货单量":
                cell.number_format = "#,##0"
            elif column == "数据发货占比":
                cell.number_format = "0.00%"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
