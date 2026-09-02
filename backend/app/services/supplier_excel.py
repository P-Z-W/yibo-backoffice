"""Excel template, import parsing, and export helpers for supplier master data."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import TypedDict

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.models.supplier import Supplier

MAX_SUPPLIER_UPLOAD_BYTES = 10 * 1024 * 1024
SUPPLIER_IMPORT_COLUMNS = [
    "序号",
    "供应商名称",
    "供应商联系人",
    "联系电话",
    "联系地址",
    "合作时间",
    "常用产品类型",
    "状态",
    "供应商备注",
    "变更说明",
]
SUPPLIER_EXPORT_COLUMNS = [
    *SUPPLIER_IMPORT_COLUMNS[:-1],
    "最后更新时间",
]


class SupplierImportRow(TypedDict):
    name: str
    contact_name: str
    contact_phone: str
    address: str
    cooperation_start_date: date | None
    product_types: str
    is_active: bool | None
    note: str | None
    change_note: str


HEADER_ALIASES = {
    "供应商名称": "name",
    "供应商": "name",
    "供应商联系人": "contact_name",
    "联系人": "contact_name",
    "联系电话": "contact_phone",
    "电话": "contact_phone",
    "手机号": "contact_phone",
    "联系地址": "address",
    "地址": "address",
    "合作时间": "cooperation_start_date",
    "合作日期": "cooperation_start_date",
    "常用产品类型": "product_types",
    "产品类型": "product_types",
    "状态": "is_active",
    "合作状态": "is_active",
    "供应商备注": "note",
    "备注": "note",
    "变更说明": "change_note",
    "本次变更说明": "change_note",
}


def _text(value: object, *, maximum: int, label: str, row_number: int) -> str:
    text = str(value or "").strip()
    if len(text) > maximum:
        raise ValueError(f"第 {row_number} 行“{label}”不能超过 {maximum} 字")
    return text


def _date(value: object, row_number: int) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError) as exc:
            raise ValueError(f"第 {row_number} 行“合作时间”不是有效日期") from exc
    normalized = str(value).strip().replace("年", "-").replace("月", "-").replace("日", "")
    normalized = normalized.replace("/", "-").replace(".", "-")
    try:
        return date.fromisoformat(normalized)
    except ValueError as exc:
        raise ValueError(f"第 {row_number} 行“合作时间”请填写为 YYYY-MM-DD") from exc


def _status(value: object, row_number: int) -> bool | None:
    text = str(value or "").strip().casefold()
    if not text:
        return None
    if text in {"合作中", "启用", "是", "active", "1", "true"}:
        return True
    if text in {"已停用", "停用", "否", "inactive", "0", "false"}:
        return False
    raise ValueError(f"第 {row_number} 行“状态”只能填写“合作中”或“已停用”")


def parse_supplier_import(contents: bytes) -> list[SupplierImportRow]:
    try:
        workbook = load_workbook(BytesIO(contents), data_only=True, read_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
    except (OSError, StopIteration, ValueError) as exc:
        raise ValueError("无法读取该 Excel 文件，请使用系统提供的导入模板") from exc

    mapping = {
        index: HEADER_ALIASES[header]
        for index, header in enumerate(headers)
        if header in HEADER_ALIASES
    }
    if "name" not in mapping.values():
        raise ValueError("表格必须包含“供应商名称”列")
    mapped_fields = set(mapping.values())

    parsed: list[SupplierImportRow] = []
    normalized_names: set[str] = set()
    errors: list[str] = []
    for row_number, values in enumerate(rows, start=2):
        raw = {
            field: values[index] if index < len(values) else None
            for index, field in mapping.items()
        }
        if not any(value not in (None, "") for value in raw.values()):
            continue
        try:
            name = _text(raw.get("name"), maximum=160, label="供应商名称", row_number=row_number)
            if not name:
                raise ValueError(f"第 {row_number} 行“供应商名称”不能为空")
            normalized_name = " ".join(name.split()).casefold()
            if normalized_name in normalized_names:
                raise ValueError(f"第 {row_number} 行供应商名称重复：{name}")
            normalized_names.add(normalized_name)
            parsed.append(
                {
                    "name": " ".join(name.split()),
                    "contact_name": _text(
                        raw.get("contact_name"),
                        maximum=100,
                        label="供应商联系人",
                        row_number=row_number,
                    ),
                    "contact_phone": _text(
                        raw.get("contact_phone"),
                        maximum=50,
                        label="联系电话",
                        row_number=row_number,
                    ),
                    "address": _text(
                        raw.get("address"),
                        maximum=255,
                        label="联系地址",
                        row_number=row_number,
                    ),
                    "cooperation_start_date": _date(
                        raw.get("cooperation_start_date"), row_number
                    ),
                    "product_types": _text(
                        raw.get("product_types"),
                        maximum=500,
                        label="常用产品类型",
                        row_number=row_number,
                    ),
                    "is_active": _status(raw.get("is_active"), row_number),
                    "note": (
                        _text(
                            raw.get("note"),
                            maximum=1000,
                            label="供应商备注",
                            row_number=row_number,
                        )
                        or None
                        if "note" in mapped_fields
                        else None
                    ),
                    "change_note": _text(
                        raw.get("change_note"),
                        maximum=500,
                        label="变更说明",
                        row_number=row_number,
                    ),
                }
            )
        except ValueError as exc:
            errors.append(str(exc))
    if errors:
        detail = "；".join(errors[:10])
        if len(errors) > 10:
            detail += f"；另有 {len(errors) - 10} 条错误"
        raise ValueError(detail)
    if not parsed:
        raise ValueError("Excel 中没有可导入的供应商数据")
    return parsed


def _style_sheet(
    sheet,
    columns: list[str],
    widths: list[int],
    last_row: int,
    *,
    body_fill_color: str = "DDEBF7",
) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(last_row, 2)}"
    sheet.sheet_view.showGridLines = True
    sheet.row_dimensions[1].height = 30
    header_fill = PatternFill("solid", fgColor="F0EDCD")
    header_border = Border(
        left=Side(style="thin", color="303030"),
        right=Side(style="thin", color="303030"),
        top=Side(style="thin", color="303030"),
        bottom=Side(style="thin", color="303030"),
    )
    body_fill = PatternFill("solid", fgColor=body_fill_color)
    thin_border = Border(
        left=Side(style="thin", color="7A8B99"),
        right=Side(style="thin", color="7A8B99"),
        top=Side(style="thin", color="7A8B99"),
        bottom=Side(style="thin", color="7A8B99"),
    )
    for column_index, (column, width) in enumerate(zip(columns, widths), start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.font = Font(bold=True, color="263F63")
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    for row_index in range(2, last_row + 1):
        sheet.row_dimensions[row_index].height = 24
        for column_index in range(1, len(columns) + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.border = thin_border
            cell.fill = body_fill
            cell.alignment = Alignment(vertical="center")


def _add_excel_table(sheet, *, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def build_supplier_template() -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "供应商导入"
    sheet.append(SUPPLIER_IMPORT_COLUMNS)
    template_last_row = 500
    _style_sheet(
        sheet,
        SUPPLIER_IMPORT_COLUMNS,
        [10, 24, 18, 18, 28, 16, 24, 14, 32, 32],
        template_last_row,
        body_fill_color="FFF2CC",
    )
    sheet["B1"].comment = Comment(
        "必填；系统按供应商名称匹配，同名更新，未匹配名称新增。", "系统"
    )
    sheet["F1"].comment = Comment("请填写日期，例如 2026/8/31。", "系统")
    sheet["G1"].comment = Comment("多个产品类型可使用逗号、顿号分隔。", "系统")
    sheet["H1"].comment = Comment("可填写“合作中”或“已停用”；新增时留空默认合作中。", "系统")
    sheet["J1"].comment = Comment("可填写本次批量维护说明，将写入月度变更记录。", "系统")
    for row_index in range(2, template_last_row + 1):
        sheet.cell(row=row_index, column=1).value = (
            f'=IF(B{row_index}="","",ROW()-1)'
        )
        sheet.cell(row=row_index, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        sheet.cell(row=row_index, column=4).number_format = "@"
        sheet.cell(row=row_index, column=6).number_format = "yyyy/m/d"
    status_validation = DataValidation(
        type="list", formula1='"合作中,已停用"', allow_blank=True
    )
    status_validation.error = "请选择“合作中”或“已停用”"
    status_validation.errorTitle = "状态填写错误"
    status_validation.prompt = "新增时留空默认为合作中；更新时留空保留原状态。"
    status_validation.promptTitle = "合作状态"
    sheet.add_data_validation(status_validation)
    status_validation.add(f"H2:H{template_last_row}")
    sheet.conditional_formatting.add(
        f"A2:J{template_last_row}",
        FormulaRule(formula=["LEN($B2)>0"], fill=PatternFill("solid", fgColor="DDEBF7")),
    )
    _add_excel_table(
        sheet,
        name="SupplierImportTable",
        ref=f"A1:J{template_last_row}",
    )
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_supplier_export(rows: list[Supplier]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "供应商档案"
    sheet.append(SUPPLIER_EXPORT_COLUMNS)
    for index, row in enumerate(rows, start=1):
        sheet.append(
            [
                index,
                row.name,
                row.contact_name or "",
                row.contact_phone or "",
                row.address or "",
                row.cooperation_start_date,
                row.product_types or "",
                "合作中" if row.is_active else "已停用",
                row.note or "",
                row.updated_at,
            ]
        )
    last_row = len(rows) + 1
    _style_sheet(
        sheet,
        SUPPLIER_EXPORT_COLUMNS,
        [10, 24, 18, 18, 28, 16, 24, 14, 32, 20],
        last_row,
    )
    for row_index in range(2, last_row + 1):
        sheet.cell(row=row_index, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        sheet.cell(row=row_index, column=4).number_format = "@"
        sheet.cell(row=row_index, column=6).number_format = "yyyy/m/d"
        sheet.cell(row=row_index, column=8).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        sheet.cell(row=row_index, column=10).number_format = "yyyy-mm-dd hh:mm"
    _add_excel_table(
        sheet,
        name="SupplierExportTable",
        ref=f"A1:J{last_row}",
    )
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
