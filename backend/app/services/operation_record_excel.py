"""Excel helpers for customer and operating-record modules."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MAX_OPERATION_UPLOAD_BYTES = 10 * 1024 * 1024


@dataclass(frozen=True)
class OperationExcelDefinition:
    code: str
    name: str
    columns: tuple[str, ...]
    fields: tuple[str, ...]
    widths: tuple[int, ...]
    required_field: str
    required_label: str


DEFINITIONS = {
    "customer_changes": OperationExcelDefinition(
        "customer_changes",
        "客户变化",
        ("记录ID", "变化类型", "发生时间", "客户名称", "来源渠道", "数量", "备注"),
        (
            "record_id",
            "change_type",
            "occurred_at",
            "customer_name",
            "source_channel",
            "quantity",
            "note",
        ),
        (12, 16, 22, 26, 20, 12, 40),
        "change_type",
        "变化类型",
    ),
    "service_issues": OperationExcelDefinition(
        "service_issues",
        "客户服务情况",
        (
            "记录ID",
            "团队名",
            "投诉大类",
            "问题详细描述",
            "核实原因",
            "责任归属",
            "整改措施",
            "状态",
        ),
        (
            "record_id",
            "team_name",
            "complaint_category",
            "issue_description",
            "verified_cause",
            "responsibility",
            "corrective_action",
            "status",
        ),
        (12, 20, 20, 38, 34, 22, 34, 16),
        "issue_description",
        "问题详细描述",
    ),
    "value_added": OperationExcelDefinition(
        "value_added",
        "增值服务",
        ("记录ID", "团队ID", "团队名称", "服务编码", "服务名称", "服务分组", "数量"),
        (
            "record_id",
            "team_id",
            "team_name",
            "service_code",
            "service_name",
            "service_group",
            "quantity",
        ),
        (12, 18, 26, 18, 26, 20, 14),
        "service_name",
        "服务名称",
    ),
    "short_video": OperationExcelDefinition(
        "short_video",
        "短视频情况",
        ("记录ID", "短视频数量", "短视频类型", "负责人", "备注"),
        ("record_id", "video_count", "video_type", "owner", "note"),
        (12, 16, 24, 18, 42),
        "video_count",
        "短视频数量",
    ),
}


def get_operation_excel_definition(dataset_type: str) -> OperationExcelDefinition:
    try:
        return DEFINITIONS[dataset_type]
    except KeyError as exc:
        raise ValueError("不支持的业务台账类型") from exc


def _style_sheet(sheet, definition: OperationExcelDefinition, last_row: int) -> None:
    sheet.freeze_panes = "A2"
    last_column = get_column_letter(len(definition.columns))
    sheet.auto_filter.ref = f"A1:{last_column}{max(last_row, 2)}"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 30
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    header_border = Border(bottom=Side(style="medium", color="4D8DF7"))
    row_border = Border(bottom=Side(style="thin", color="E2E9F3"))
    for index, (column, width) in enumerate(
        zip(definition.columns, definition.widths), start=1
    ):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(bold=True, color="263F63")
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row_index in range(2, last_row + 1):
        sheet.row_dimensions[row_index].height = 24
        for column_index in range(1, len(definition.columns) + 1):
            sheet.cell(row=row_index, column=column_index).border = row_border
            sheet.cell(row=row_index, column=column_index).alignment = Alignment(
                vertical="center"
            )


def build_operation_template(dataset_type: str) -> BytesIO:
    definition = get_operation_excel_definition(dataset_type)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = definition.name[:31]
    sheet.append(list(definition.columns))
    _style_sheet(sheet, definition, 2)
    sheet["A1"].comment = Comment(
        "新增记录请留空；更新已有记录时填写系统导出文件中的记录ID。", "系统"
    )
    if dataset_type == "customer_changes":
        validation = DataValidation(type="list", formula1='"新进,流失,意向"')
        validation.error = "请选择新进、流失或意向"
        sheet.add_data_validation(validation)
        validation.add("B2:B500")
        sheet["B1"].comment = Comment("必填；请选择新进、流失或意向。", "系统")
        sheet["C1"].comment = Comment("保留完整时间点；月度统计将按该时间归属。", "系统")
        sheet["F1"].comment = Comment("留空时按 1 计。", "系统")
        sheet["F2"].number_format = "#,##0"
    elif dataset_type == "service_issues":
        validation = DataValidation(
            type="list", formula1='"待核实,整改中,已完成,已关闭"', allow_blank=True
        )
        validation.error = "请选择待核实、整改中、已完成或已关闭"
        sheet.add_data_validation(validation)
        validation.add("H2:H500")
        sheet["D1"].comment = Comment("必填；请描述具体问题。", "系统")
    elif dataset_type == "value_added":
        sheet["C1"].comment = Comment("必填；请输入团队名称。", "系统")
        sheet["E1"].comment = Comment("必填；请输入服务名称。", "系统")
        sheet["G1"].comment = Comment("必填；请输入非负整数。", "系统")
        sheet["G2"].number_format = "#,##0"
    else:
        sheet["B1"].comment = Comment("必填；请输入非负整数。", "系统")
        sheet["B2"].number_format = "#,##0"
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def parse_operation_import(
    dataset_type: str, contents: bytes
) -> list[dict[str, object]]:
    definition = get_operation_excel_definition(dataset_type)
    try:
        workbook = load_workbook(BytesIO(contents), data_only=True, read_only=True)
        rows = workbook.active.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
    except (OSError, StopIteration, ValueError) as exc:
        raise ValueError("无法读取该 Excel 文件，请使用系统提供的导入模板") from exc
    header_fields = dict(zip(definition.columns, definition.fields))
    mapping = {
        index: header_fields[header]
        for index, header in enumerate(headers)
        if header in header_fields
    }
    if definition.required_field not in mapping.values():
        raise ValueError(f"表格必须包含“{definition.required_label}”列")

    parsed: list[dict[str, object]] = []
    record_ids: set[int] = set()
    for row_number, values in enumerate(rows, start=2):
        raw = {
            field: values[index] if index < len(values) else None
            for index, field in mapping.items()
        }
        if not any(value not in (None, "") for value in raw.values()):
            continue
        record_id_value = raw.get("record_id")
        record_id = None
        if record_id_value not in (None, ""):
            try:
                record_id = int(record_id_value)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"第 {row_number} 行“记录ID”必须是整数") from exc
            if record_id <= 0 or record_id in record_ids:
                raise ValueError(f"第 {row_number} 行“记录ID”无效或重复")
            record_ids.add(record_id)
        row = {
            field: value
            for field, value in raw.items()
            if field != "record_id" and value is not None
        }
        row["record_id"] = record_id
        row["_excel_row"] = row_number
        if dataset_type == "customer_changes" and row.get("quantity") in (None, ""):
            row["quantity"] = 1
        if dataset_type == "service_issues" and not str(row.get("status") or "").strip():
            row["status"] = "待核实"
        parsed.append(row)
    if not parsed:
        raise ValueError("Excel 中没有可导入的数据")
    return parsed


def build_operation_export(
    dataset_type: str, rows: list[dict[str, object]]
) -> BytesIO:
    definition = get_operation_excel_definition(dataset_type)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = definition.name[:31]
    sheet.append(list(definition.columns))
    for row in rows:
        sheet.append([row.get(field) for field in definition.fields])
    last_row = len(rows) + 1
    _style_sheet(sheet, definition, last_row)
    for row_index in range(2, last_row + 1):
        sheet.cell(row=row_index, column=1).number_format = "0"
        if dataset_type == "customer_changes":
            sheet.cell(row=row_index, column=3).number_format = "yyyy-mm-dd hh:mm:ss"
            sheet.cell(row=row_index, column=6).number_format = "#,##0"
        elif dataset_type == "value_added":
            sheet.cell(row=row_index, column=7).number_format = "#,##0"
        elif dataset_type == "short_video":
            sheet.cell(row=row_index, column=2).number_format = "#,##0"
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
